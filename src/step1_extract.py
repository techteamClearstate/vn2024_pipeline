"""
Step 1 — Extraction
===================
1. Stream the large VN .xlsx (520k rows) to a flat TSV cache via openpyxl
   read-only mode (avoids loading the whole workbook into memory).
2. Build the keyword lookup + prefix-trie from the V0 reference 'Updated' sheet.
"""
import csv
import pickle
import re
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import load_workbook

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import settings as cfg


def _require(path, label):
    """Boundary check: fail clearly if an expected input file is absent."""
    if not Path(path).exists():
        raise FileNotFoundError(
            f"{label} not found: {path}\n"
            f"Place the source workbooks in {cfg.UPLOADS_DIR} (see README).")


def _require_columns(have, need, where):
    """Boundary check: fail clearly listing any expected columns that are
    missing, instead of a cryptic KeyError deep in the pipeline."""
    missing = [c for c in need if c not in have]
    if missing:
        raise ValueError(
            f"{where}: missing expected column(s) {missing}.\n"
            f"Found columns: {list(have)}\n"
            f"Fix the source file or update the names in config/settings.py.")


def extract_vn_to_tsv() -> int:
    """Stream the source import file to a TSV cache. Returns row count.

    Accepts either an .xlsx workbook (streamed sheet-by-row via openpyxl) or a
    .csv export (streamed row-by-row) — some markets ship processed CSVs rather
    than workbooks. Both produce the same TSV cache the rest of the pipeline
    reads, so only the reader differs.
    """
    cfg.INTERMEDIATE.mkdir(parents=True, exist_ok=True)
    _require(cfg.VN_SOURCE_XLSX, "source import file")
    src = Path(cfg.VN_SOURCE_XLSX)
    if src.suffix.lower() == ".csv":
        return _extract_csv_to_tsv(src)
    return _extract_xlsx_to_tsv(src)


def _extract_csv_to_tsv(src: Path) -> int:
    """Stream a .csv source (e.g. a market's processed import extract) to the TSV
    cache, validating the header. Tries utf-8 then latin-1 (trade CSVs vary)."""
    need = [cfg.VN_DESCRIPTION_COL, cfg.VN_HS4_COL, cfg.VN_HS_CODE_COL,
            *cfg.MANUFACTURER_PARTY_COLS]
    for enc in ("utf-8-sig", "latin-1"):
        try:
            with open(src, "r", newline="", encoding=enc) as fin:
                reader = csv.reader(fin)
                header = next(reader, [])
                _require_columns(header, need, f"CSV source '{src.name}'")
                n = 0
                with open(cfg.VN_TSV, "w", newline="", encoding="utf-8") as fout:
                    writer = csv.writer(fout, delimiter="\t")
                    writer.writerow(header)
                    for row in reader:
                        writer.writerow(row)
                        n += 1
                        if n % 50000 == 0:
                            print(f"    ...{n:,} rows streamed", flush=True)
            print(f"  [extract] wrote {n:,} rows to {cfg.VN_TSV.name} (from CSV)")
            return n
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode CSV source {src} as utf-8 or latin-1.")


def _extract_xlsx_to_tsv(src: Path) -> int:
    """Stream the RawData sheet of an .xlsx workbook to the TSV cache."""
    wb = load_workbook(src, read_only=True, data_only=True)
    if cfg.VN_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"VN source workbook has no sheet '{cfg.VN_SHEET}'. "
            f"Available sheets: {wb.sheetnames}. "
            f"Update cfg.VN_SHEET in config/settings.py.")
    ws = wb[cfg.VN_SHEET]

    header = next(ws.iter_rows(values_only=True), ())
    need = [cfg.VN_DESCRIPTION_COL, cfg.VN_HS4_COL, cfg.VN_HS_CODE_COL,
            *cfg.MANUFACTURER_PARTY_COLS]
    _require_columns(header, need, f"VN sheet '{cfg.VN_SHEET}'")

    n = 0
    with open(cfg.VN_TSV, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh, delimiter="\t")
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
            n += 1
            if n % 50000 == 0:
                print(f"    ...{n:,} rows streamed", flush=True)
    wb.close()
    print(f"  [extract] wrote {n:,} rows to {cfg.VN_TSV.name}")
    return n


def _load_v0_reference() -> pd.DataFrame:
    """Load + validate the reference sheet (shared by lookup + lexicon).

    The active reference (master "List of companies by sub-OU") carries its
    header on cfg.V0_HEADER_ROW and names its columns differently, so we read at
    that header row and rename via cfg.V0_SOURCE_COLS to the logical V0_COLS
    names the rest of the pipeline expects. Both are optional (None → default
    header / no rename) so a V0-formatted sheet still loads unchanged.
    """
    _require(cfg.V0_REFERENCE_XLSX, "V0 reference workbook")
    header = getattr(cfg, "V0_HEADER_ROW", None)
    try:
        v0 = pd.read_excel(cfg.V0_REFERENCE_XLSX, sheet_name=cfg.V0_SHEET,
                           header=header if header is not None else 0)
    except ValueError as e:
        raise ValueError(
            f"reference workbook has no sheet '{cfg.V0_SHEET}'. "
            f"Update cfg.V0_SHEET in config/settings.py.") from e

    source_map = getattr(cfg, "V0_SOURCE_COLS", None)
    if source_map:
        _require_columns(v0.columns, source_map.keys(),
                         f"reference sheet '{cfg.V0_SHEET}'")
        v0 = v0.rename(columns=source_map)
    _require_columns(v0.columns, cfg.V0_COLS.values(),
                     f"reference sheet '{cfg.V0_SHEET}'")
    return v0


def _load_master_reference() -> pd.DataFrame:
    """Load the full master sheet, including generic-family rows.

    The lookup/lexicon builders intentionally use cfg.V0_SHEET
    ("Updated (excl. generic)") so generic labels do not become match keywords.
    The final reference gate needs the full "Updated" sheet to distinguish strict
    master rows from generic-family rows that should be parked for review.
    """
    _require(cfg.V0_REFERENCE_XLSX, "V0 reference workbook")
    sheet = getattr(cfg, "V0_MASTER_SHEET", "Updated")
    header = getattr(cfg, "V0_HEADER_ROW", None)
    try:
        v0 = pd.read_excel(cfg.V0_REFERENCE_XLSX, sheet_name=sheet,
                           header=header if header is not None else 0)
    except ValueError as e:
        raise ValueError(
            f"reference workbook has no sheet '{sheet}'. "
            f"Update cfg.V0_MASTER_SHEET in config/settings.py.") from e

    source_map = getattr(cfg, "V0_SOURCE_COLS", None)
    if source_map:
        _require_columns(v0.columns, source_map.keys(),
                         f"reference sheet '{sheet}'")
        v0 = v0.rename(columns=source_map)
    need = [*cfg.V0_COLS.values(), "Generic Family Name?"]
    _require_columns(v0.columns, need, f"reference sheet '{sheet}'")
    return v0


def build_keyword_lookup() -> int:
    """Build keyword → mapping dict and 4-char prefix trie from V0 reference."""
    v0 = _load_v0_reference()
    c = cfg.V0_COLS
    ref = v0[[c["segment"], c["sub_segment"], c["product"],
              c["player"], c["keyword"]]].dropna(subset=[c["keyword"]]).copy()

    ref[c["keyword"]] = ref[c["keyword"]].astype(str).str.strip()
    # Drop pure-numeric and too-short keywords
    ref = ref[~ref[c["keyword"]].str.match(r"^\d+\.?\d*$")]
    ref = ref[ref[c["keyword"]].str.len() >= cfg.MIN_KEYWORD_LEN]
    # Longest-first so specific keywords win; dedupe on keyword
    ref = ref.sort_values(c["keyword"], key=lambda s: s.str.len(), ascending=False)
    ref = ref.drop_duplicates(subset=[c["keyword"]], keep="first")

    lookup = {}
    for _, r in ref.iterrows():
        kw = r[c["keyword"]].lower()
        if kw in cfg.BLACKLIST:          # apply blacklist
            continue
        lookup[kw] = {
            "Segment":     str(r[c["segment"]])     if pd.notna(r[c["segment"]])     else "",
            "Sub-segment": str(r[c["sub_segment"]]) if pd.notna(r[c["sub_segment"]]) else "",
            "Product":     str(r[c["product"]])     if pd.notna(r[c["product"]])     else "",
            "Player":      str(r[c["player"]])      if pd.notna(r[c["player"]])      else "",
            "Family_Name": r[c["keyword"]],
        }

    # Benchmark-supervised gap-fill: merge brand keywords mined from the human
    # labels (train split only). The reference always WINS on a key conflict, so
    # harvested entries only ADD families the curated reference lacked.
    n_harvest = _merge_harvested_keywords(lookup)

    prefix_map = defaultdict(list)
    for kw in sorted(lookup.keys(), key=len, reverse=True):
        if len(kw) >= cfg.PREFIX_LEN:
            prefix_map[kw[:cfg.PREFIX_LEN]].append(kw)

    with open(cfg.V0_LOOKUP_PKL, "wb") as fh:
        pickle.dump(lookup, fh)
    with open(cfg.PREFIX_MAP_PKL, "wb") as fh:
        pickle.dump(dict(prefix_map), fh)

    extra = f" (+{n_harvest:,} benchmark-harvested)" if n_harvest else ""
    print(f"  [lookup] {len(lookup):,} active keywords "
          f"({len(prefix_map):,} prefix buckets) after blacklist{extra}")
    return len(lookup)


def _merge_harvested_keywords(lookup: dict) -> int:
    """Add benchmark-harvested brand keywords into `lookup` in place, skipping any
    keyword the reference already defines (reference wins) or that the blacklist
    forbids. Returns the count of net-new keywords added. No-op unless
    cfg.USE_BENCHMARK_HARVEST and the harvest pickle exists."""
    if not getattr(cfg, "USE_BENCHMARK_HARVEST", False):
        return 0
    path = getattr(cfg, "HARVEST_KEYWORDS_PKL", None)
    if not path or not Path(path).exists():
        return 0
    with open(path, "rb") as fh:
        harvested = pickle.load(fh)
    added = 0
    for kw, rec in harvested.items():
        if kw in lookup or kw in cfg.BLACKLIST or len(kw) < cfg.MIN_KEYWORD_LEN:
            continue
        lookup[kw] = rec
        added += 1
    return added


def _canon_sep(s) -> str:
    """Normalize a product label's SEPARATORS/whitespace without reordering words:
    unify ' | ' and spaced ' - ' to '_', tidy spaces around '_', collapse runs.
    Intra-word hyphens (e.g. "Non-Absorbable") are left intact — only a hyphen
    flanked by spaces is treated as a Head/Qualifier separator."""
    s = str(s)
    s = re.sub(r"\s*\|\s*", "_", s)     # pipe separator  → underscore
    s = re.sub(r"\s+-\s+", "_", s)      # spaced hyphen   → underscore
    s = re.sub(r"\s*_\s*", "_", s)      # tidy spaces around underscore
    return re.sub(r"\s+", " ", s).strip()


def _canon_tokens(s) -> frozenset:
    """Case/separator/order-insensitive identity of a product label."""
    return frozenset(t for t in re.split(r"[^a-z0-9]+", str(s).lower()) if t)


def _choose_canonical(variants) -> str:
    """Pick one canonical label for a set of same-token-set variants: prefer the
    "Head_Qualifier" underscore form, then the shorter, then alphabetical (which
    favours Title-case over lower-case). Deterministic."""
    cands = sorted({_canon_sep(v) for v in variants},
                   key=lambda x: ("_" not in x, len(x), x))
    return cands[0]


def build_product_canonical_map() -> int:
    """Build {raw Product label → canonical label} from the reference so every
    spelling/separator/order variant of one product collapses to a single label.
    Applied to Product_V0 in step3 so the Dashboard lists each product once."""
    v0 = _load_v0_reference()
    labels = [str(x) for x in v0[cfg.V0_COLS["product"]].dropna().unique()]
    groups = defaultdict(list)
    for lbl in labels:
        groups[_canon_tokens(lbl)].append(lbl)

    cmap = {}
    for variants in groups.values():
        canon = _choose_canonical(variants)
        for lbl in variants:
            if lbl != canon:
                cmap[lbl] = canon

    with open(cfg.PRODUCT_CANONICAL_PKL, "wb") as fh:
        pickle.dump(cmap, fh)
    print(f"  [canon] {len(cmap):,} product-label variants mapped to "
          f"{len({*cmap.values()}):,} canonical labels")
    return len(cmap)


def canonicalize_products(series: pd.Series) -> pd.Series:
    """Apply the persisted canonical map to a Product label column (no-op if the
    map is absent). Blanks pass through unchanged."""
    if not cfg.PRODUCT_CANONICAL_PKL.exists():
        return series
    with open(cfg.PRODUCT_CANONICAL_PKL, "rb") as fh:
        cmap = pickle.load(fh)
    return series.map(lambda x: cmap.get(x, x) if isinstance(x, str) else x)


def norm_exact(s) -> str:
    """Normalize a taxonomy dimension (Segment / Sub-segment / Product) for the
    reference-tuple gate: string, strip, collapse whitespace, casefold. Applied
    identically to the reference and the mapped output so they compare like-for-like."""
    return re.sub(r"\s+", " ", str(s if s is not None else "")).strip().casefold()


def norm_loose(s) -> str:
    """Reference-label normalization that also folds common separators and marks.

    Used only for relabelling to the exact master wording after a punctuation or
    spacing mismatch, e.g. Product_Qualifier vs Product - Qualifier.
    """
    t = str(s if s is not None else "")
    t = re.sub(r"[™®©]", "", t)
    t = re.sub(r"[_\-–—/\\]+", " ", t)
    return re.sub(r"\s+", " ", t).strip().casefold()


def _norm_dim(s) -> str:
    return norm_exact(s)


def build_reference_tuples() -> int:
    """Build the final reference-compliance lookup structures from the master.

    Strict family rows are validated on the full five-key
    (Segment, Sub-segment, Product, Player, Model/Family). Generic-family rows
    are indexed separately so they can be marked review-only. Category and loose
    maps support controlled relabelling to the master's exact wording.

    MUST run after build_product_canonical_map so canonicalize_products is a no-op-
    free, faithful match of what step3 writes into Product_V0.
    """
    v0 = _load_master_reference().fillna("")
    c = cfg.V0_COLS

    ref = v0[[c["segment"], c["sub_segment"], c["product"],
              c["player"], c["keyword"], "Generic Family Name?"]].copy()
    ref[c["product"]] = canonicalize_products(ref[c["product"]].astype(str))
    gen_flag = ref["Generic Family Name?"].astype(str).str.strip()
    strict = ref[gen_flag == ""]
    generic = ref[gen_flag != ""]

    def trip(row):
        return (str(row[c["segment"]]), str(row[c["sub_segment"]]),
                str(row[c["product"]]))

    def full(row):
        return (*trip(row), str(row[c["player"]]), str(row[c["keyword"]]))

    category_exact, products = set(), set()
    category_votes = defaultdict(Counter)
    pf_cats = defaultdict(set)

    for _, row in ref.iterrows():
        t = trip(row)
        pn = norm_exact(t[2])
        if not pn or pn == norm_exact(cfg.UNSPECIFIED_LABEL):
            continue
        category_exact.add(tuple(norm_exact(x) for x in t))
        category_votes[tuple(norm_loose(x) for x in t)][t] += 1
        pf_cats[(norm_loose(row[c["player"]]),
                 norm_loose(row[c["keyword"]]))].add(t)
        products.add(pn)

    full_exact, full_loose = set(), {}
    generic_exact, generic_loose = {}, {}

    for _, row in strict.iterrows():
        f = full(row)
        if not norm_exact(f[4]):
            continue
        full_exact.add(tuple(norm_exact(x) for x in f))
        full_loose.setdefault(tuple(norm_loose(x) for x in f), f)

    for _, row in generic.iterrows():
        f = full(row)
        if not norm_exact(f[4]):
            continue
        payload = (f, str(row["Generic Family Name?"]))
        generic_exact.setdefault(tuple(norm_exact(x) for x in f), payload)
        generic_loose.setdefault(tuple(norm_loose(x) for x in f), payload)

    category_loose = {
        k: v.most_common(1)[0][0]
        for k, v in category_votes.items()
    }
    pf_cats = {k: sorted(v) for k, v in pf_cats.items()}

    payload = {
        "full_exact": full_exact,
        "full_loose": full_loose,
        "generic_exact": generic_exact,
        "generic_loose": generic_loose,
        "gen_exact": generic_exact,        # compatibility with the compliance tool
        "gen_loose": generic_loose,
        "category_exact": category_exact,
        "category_loose": category_loose,
        "cat_exact": category_exact,
        "cat_loose": category_loose,
        "pf_cats": pf_cats,
        "triples": category_exact,         # backward-compatible category keyset
        "products": products,
        "n_strict": len(strict),
        "n_generic": len(generic),
        "n_all": len(ref),
    }
    with open(cfg.REFERENCE_TUPLES_PKL, "wb") as fh:
        pickle.dump(payload, fh)
    print(f"  [ref-gate] {len(full_exact):,} strict full keys / "
          f"{len(generic_exact):,} generic full keys / "
          f"{len(category_exact):,} category tuples from the master taxonomy")
    return len(full_exact)


def norm_phrase(s) -> str:
    """Normalize text for category-phrase matching (used on both the lexicon
    and the description so they line up): lowercase, split compound separators,
    crude singularize, strip punctuation, collapse whitespace."""
    s = str(s).lower()
    s = re.sub(r"[_/]", " ", s)
    s = re.sub(r"s\b", "", s)          # crude singularize (symmetric both sides)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def build_category_lexicon() -> int:
    """Build the Tier-2 category phrase lexicon from the V0 reference.

    Two phrase sources, both keyed by normalized phrase:
      * curated qualifier→Product map        → confidence "high"
      * multi-word reference Product labels   → confidence "med"
    Segment/Sub-segment for each phrase are resolved as the most-frequent pair
    for that Product in the reference.
    """
    v0 = _load_v0_reference()
    c = cfg.V0_COLS
    ref = v0.dropna(subset=[c["product"]]).copy()

    # Most-frequent (Segment, Sub-segment) per canonical Product label.
    def resolve(product):
        rows = ref[ref[c["product"]] == product]
        if rows.empty:
            return None
        key = rows.groupby([c["segment"], c["sub_segment"]], dropna=False).size()
        seg, sub = key.idxmax()
        return {
            "Segment":     str(seg) if pd.notna(seg) else "",
            "Sub-segment": str(sub) if pd.notna(sub) else "",
            "Product":     product,
        }

    lex = {}
    # Source 1 — curated qualifier map (high confidence)
    missing = []
    for phrase, product in cfg.CATEGORY_QUALIFIER_MAP.items():
        rec = resolve(product)
        if rec is None:
            missing.append(product)
            continue
        rec = dict(rec, confidence="high")
        lex[norm_phrase(phrase)] = rec
    if missing:
        print(f"  [category] WARNING qualifier-map Products not in reference: "
              f"{sorted(set(missing))}")

    # Source 2 — multi-word reference Product labels (med confidence)
    for product in ref[c["product"]].dropna().astype(str).unique():
        phrase = norm_phrase(product)
        if len(phrase.split()) < 2 or len(phrase) < 6:
            continue                                   # skip single-token/abbrev
        if phrase in cfg.GENERIC_LABEL_BLACKLIST:
            continue
        if phrase in lex:                              # qualifier map wins
            continue
        rec = resolve(product)
        if rec:
            lex[phrase] = dict(rec, confidence="med")

    with open(cfg.CATEGORY_LEX_PKL, "wb") as fh:
        pickle.dump(lex, fh)
    print(f"  [category] {len(lex):,} category phrases "
          f"({sum(1 for v in lex.values() if v['confidence']=='high')} high / "
          f"{sum(1 for v in lex.values() if v['confidence']=='med')} med)")
    return len(lex)


def norm_party(s) -> str:
    """Normalize a trade-party (Importer/Exporter) string for manufacturer-core
    matching: lowercase, drop '&'/punctuation, collapse whitespace."""
    s = str(s).lower()
    s = re.sub(r"&", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def build_manufacturer_lexicon() -> int:
    """Build the Tier-3 alias lexicon: distinctive 'core' phrase → canonical
    manufacturer, emitted as a longest-first list so the most-specific core
    wins. Sourced from the curated cfg.MANUFACTURER_ALIASES (no auto-derivation:
    single generic tokens collide across companies and mislabel the maker)."""
    core_to_mfr = {}
    for canonical, cores in cfg.MANUFACTURER_ALIASES.items():
        for core in cores:
            core_to_mfr[norm_party(core)] = canonical
    n_curated = len(core_to_mfr)

    # Benchmark-supervised gap-fill: add maker-name aliases mined from the human
    # labels (train split only). Curated cores WIN on conflict.
    n_harvest = 0
    if getattr(cfg, "USE_BENCHMARK_HARVEST", False):
        path = getattr(cfg, "HARVEST_MANUFACTURERS_PKL", None)
        if path and Path(path).exists():
            with open(path, "rb") as fh:
                for core, canonical in pickle.load(fh):
                    core = norm_party(core)
                    if core and core not in core_to_mfr:
                        core_to_mfr[core] = canonical
                        n_harvest += 1

    ordered = sorted(core_to_mfr.items(), key=lambda kv: len(kv[0]), reverse=True)

    with open(cfg.MANUFACTURER_ALIAS_PKL, "wb") as fh:
        pickle.dump(ordered, fh)
    extra = f" (+{n_harvest:,} benchmark-harvested)" if n_harvest else ""
    print(f"  [manufacturer] {len(ordered):,} alias cores → "
          f"{len(set(core_to_mfr.values())):,} manufacturers{extra}")
    return len(ordered)


if __name__ == "__main__":
    print("Step 1 — Extraction")
    extract_vn_to_tsv()
    build_keyword_lookup()
    build_category_lexicon()
    build_manufacturer_lexicon()
    build_product_canonical_map()
    build_reference_tuples()
