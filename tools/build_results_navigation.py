"""Build the governed, static results-navigation hub.

The site is a read-only view over the prediction-audit SQLite authority.  It
does not modify mapped workbooks, reference data, reviewer decisions, or the
production dashboard.  Publish the generated directory as one linked set.
"""

from __future__ import annotations

import argparse
import html
import json
import sqlite3
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from build_funnel_dashboard import _build_simulator


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = ROOT / "outputs/20260713_llm_adjudication/raw_outputs/prediction_audit.sqlite"
DEFAULT_OUT = ROOT / "outputs/20260713_llm_adjudication/dashboard/site"
DEFAULT_WB_DIR = ROOT / "outputs/20260713_llm_adjudication/raw_outputs/mapped_results"
DEFAULT_REVIEW_DB = ROOT / "outputs/20260713_llm_adjudication/raw_outputs/llm_review_authority.sqlite"
DIMENSIONS = {
    "segment": "segment",
    "sub_segment": "sub_segment",
    "product": "product",
    "manufacturer": "manufacturer",
    "family": "family",
}
MACRO = {
    "Vietnam": {"population": 101_598_527, "gdp": 514_697_000_000.0},
    "India": {"population": 1_463_865_525, "gdp": 3_956_067_000_000.0},
    "Pakistan": {"population": 255_219_554, "gdp": 407_307_000_000.0},
}
WB_SOURCE = "https://data.worldbank.org/?locations=IN-PK-VN"
SITE_AUDIT_LABEL = "current audit"


def esc(value: object) -> str:
    return html.escape(str(value), quote=True)


def fmt_money(value: float) -> str:
    value = float(value or 0)
    if abs(value) >= 1_000_000_000:
        return f"${value / 1_000_000_000:,.2f}B"
    if abs(value) >= 1_000_000:
        return f"${value / 1_000_000:,.1f}M"
    return f"${value:,.0f}"


def fmt_number(value: float) -> str:
    return f"{float(value or 0):,.0f}"


def aggregate(conn: sqlite3.Connection) -> tuple[list[dict], list[dict]]:
    totals = []
    for row in conn.execute(
        """SELECT output_file_id,country,fiscal_year,output_tier,count(*),
                  coalesce(sum(value_usd),0),coalesce(sum(volume),0)
             FROM row_fact GROUP BY 1,2,3,4 ORDER BY 1,4"""
    ):
        totals.append(dict(zip(
            ("file", "country", "year", "tier", "rows", "value", "volume"), row
        )))

    detail = []
    for key, column in DIMENSIONS.items():
        sql = f"""SELECT output_file_id,country,fiscal_year,output_tier,
                         coalesce(nullif(trim({column}),''),'<Unmapped>'),
                         count(*),coalesce(sum(value_usd),0),coalesce(sum(volume),0)
                    FROM row_fact GROUP BY 1,2,3,4,5 ORDER BY 1,4,7 DESC"""
        for row in conn.execute(sql):
            detail.append(dict(zip(
                ("file", "country", "year", "tier", "name", "rows", "value", "volume"), row
            )) | {"dimension": key})
    return totals, detail


def workbook_schemas(workbook_dir: Path) -> list[dict]:
    schemas = []
    for path in sorted(workbook_dir.glob("*_ML_Map_Mapped.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=False)
        sheets = []
        for ws in wb.worksheets:
            header = []
            if ws.max_row:
                header = [str(v) if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            sheets.append({"name": ws.title, "rows": ws.max_row, "columns": header})
        schemas.append({"file": path.name, "sheets": sheets})
        wb.close()
    return schemas


def load_scorecard(path: Path | None) -> dict:
    """Load aggregate-only comparison metrics; never expose row-level audit data."""
    if path is None or not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def load_review_summary(path: Path | None) -> dict:
    if path is None or not path.exists():
        return {}
    con = sqlite3.connect(f"file:{path.resolve().as_posix()}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    try:
        totals = dict(con.execute(
            """SELECT COUNT(*) reviewed,
                      SUM(CASE WHEN f.final_decision='APPROVE' THEN 1 ELSE 0 END) approved,
                      COALESCE(SUM(CASE WHEN f.final_decision='APPROVE' THEN r.cluster_value_usd END),0) approved_value
                 FROM proposal_raw r JOIN proposal_final f USING(proposal_id)"""
        ).fetchone())
        top = [dict(r) for r in con.execute(
            """SELECT r.alias_term, r.proposed_player, r.proposed_family,
                      r.proposal_type, r.cluster_rows, r.cluster_value_usd
                 FROM proposal_raw r JOIN proposal_final f USING(proposal_id)
                WHERE f.final_decision='APPROVE'
                ORDER BY r.cluster_value_usd DESC LIMIT 8"""
        )]
        precision = [dict(r) for r in con.execute(
            """SELECT p.output_tier tier, COUNT(*) sample_rows,
                      AVG(c.exact_agreement) exact_agreement
                 FROM precision_raw p JOIN precision_consensus c USING(precision_id)
                WHERE p.sample_type='Deterministic stratified random'
                GROUP BY p.output_tier ORDER BY p.output_tier"""
        )]
    finally:
        con.close()
    return {"totals": totals, "top_approved": top, "precision_agreement": precision}


def shell(title: str, active: str, body: str, script: str = "") -> str:
    nav = [("index.html", "Overview"), ("quality.html", "Weekend scorecard"), ("simulator.html", "What-if playground"), ("comparison.html", "Compare results"),
           ("outputs.html", "Outputs & tracking"), ("schemas.html", "Data schemas")]
    links = "".join(f'<a class="{("active" if label == active else "")}" href="{href}">{label}</a>' for href, label in nav)
    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>{esc(title)}</title>
<link rel="stylesheet" href="assets/site.css"></head><body>
<header><div class="brand"><div><b>Clearstate</b><small>Surgical mapping results</small></div></div>
<nav aria-label="Primary">{links}</nav></header><main>{body}</main>
<footer>Aggregate-only dashboard · Authority: {esc(SITE_AUDIT_LABEL)} · Generated {date.today().isoformat()} · <a href="{WB_SOURCE}" target="_blank" rel="noopener">2025 population and GDP: World Bank</a></footer>
<script src="assets/data.js"></script><script src="assets/site.js"></script>{script}</body></html>"""


def write_site(
    out: Path,
    totals: list[dict],
    detail: list[dict],
    schemas: list[dict],
    scorecard: dict,
    review: dict,
    audit_label: str,
    raw_link_prefix: str,
    simulator: dict,
) -> None:
    global SITE_AUDIT_LABEL
    SITE_AUDIT_LABEL = audit_label
    assets = out / "assets"
    assets.mkdir(parents=True, exist_ok=True)
    packed_detail = [[x[k] for k in ("file", "country", "year", "tier", "name",
                                     "rows", "value", "volume", "dimension")]
                     for x in detail]
    payload = {"totals": totals, "detail": packed_detail, "schemas": schemas, "macro": MACRO,
               "simulator": simulator,
               "scorecard": scorecard, "review": review,
               "metadata": {"audit": audit_label, "macro_year": 2025,
                            "world_bank": WB_SOURCE, "detail_fields":
                            ["file", "country", "year", "tier", "name", "rows", "value", "volume", "dimension"]}}
    unpack = ";RESULTS_DATA.detail=RESULTS_DATA.detail.map(r=>Object.fromEntries(RESULTS_DATA.metadata.detail_fields.map((k,i)=>[k,r[i]])));"
    (assets / "data.js").write_text("window.RESULTS_DATA=" + json.dumps(payload, separators=(",", ":")) + unpack, encoding="utf-8")

    css = """:root{--ink:#111827;--muted:#667085;--blue:#0047ab;--line:#d9dee7;--bg:#f5f7fa;--good:#067647;--warn:#b54708;--bad:#b42318;--r:7px}*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--ink);font:14px/1.45 Inter,Segoe UI,Arial,sans-serif}header{background:#fff;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;padding:13px max(20px,calc((100% - 1440px)/2));gap:24px}.brand{display:flex;align-items:center;gap:10px;white-space:nowrap}.brand small{display:block;color:var(--muted)}.mark{display:grid;place-items:center;width:36px;height:36px;background:var(--blue);color:#fff;border-radius:6px;font-weight:800}nav{display:flex;gap:4px;flex-wrap:wrap}nav a{color:#344054;text-decoration:none;padding:8px 10px;border-radius:5px}nav a:hover,nav a.active{background:#eef4ff;color:var(--blue)}main{max-width:1440px;margin:auto;padding:26px 20px 50px}.eyebrow{color:var(--blue);font-weight:700;text-transform:uppercase;letter-spacing:.08em;font-size:11px}h1{font-size:30px;line-height:1.15;margin:6px 0 8px}h2{font-size:20px;margin:28px 0 12px}h3{font-size:15px;margin:0 0 7px}p{margin:6px 0 12px}.muted{color:var(--muted)}.notice{background:#fff7ed;border-left:4px solid var(--warn);padding:12px 14px;margin:18px 0}.grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:12px}.card{background:#fff;border:1px solid var(--line);border-radius:var(--r);padding:15px}.card a{color:var(--blue);font-weight:650}.kpi{font-size:25px;font-weight:750}.toolbar{display:flex;gap:10px;align-items:end;flex-wrap:wrap;background:#fff;border:1px solid var(--line);padding:12px;border-radius:var(--r);position:sticky;top:0;z-index:2}.toolbar label{display:grid;gap:4px;color:var(--muted);font-size:12px}.toolbar select,.toolbar input{font:inherit;padding:7px 9px;border:1px solid #aeb7c5;border-radius:5px;background:#fff;min-width:135px}.table-wrap{overflow:auto;background:#fff;border:1px solid var(--line);border-radius:var(--r);margin-top:12px}table{border-collapse:collapse;width:100%}th,td{text-align:left;padding:9px 11px;border-bottom:1px solid #e7eaf0;white-space:nowrap}th{position:sticky;top:0;background:#f9fafb;color:#475467;font-size:12px}td.num,th.num{text-align:right;font-variant-numeric:tabular-nums}.bar{height:7px;background:#e8edf5;border-radius:5px;overflow:hidden;min-width:90px}.bar i{display:block;height:100%;background:var(--blue)}.status{display:inline-block;padding:2px 7px;border-radius:20px;font-size:12px;background:#eef4ff;color:var(--blue)}.flag-low{color:var(--bad);font-weight:700}.flag-mid{color:var(--warn);font-weight:700}.flag-ok{color:var(--good);font-weight:700}.schema details{background:#fff;border:1px solid var(--line);border-radius:var(--r);margin:8px 0}.schema summary{cursor:pointer;padding:12px 14px;font-weight:650}.schema .inside{padding:0 14px 14px}.columns{display:flex;flex-wrap:wrap;gap:5px}.columns code{background:#f2f4f7;padding:3px 6px;border-radius:4px;font-size:11px}.destination{display:flex;justify-content:space-between;gap:16px;align-items:center}.destination p{margin:2px 0}.links a{display:block;margin:5px 0}footer{max-width:1440px;margin:auto;padding:15px 20px 35px;color:var(--muted);border-top:1px solid var(--line)}@media(max-width:800px){header{align-items:flex-start;flex-direction:column}.grid{grid-template-columns:1fr}.toolbar{position:static}.destination{align-items:flex-start;flex-direction:column}h1{font-size:25px}}"""
    css += ".gates{display:flex;gap:8px;flex-wrap:wrap}.toolbar .gatecheck{display:flex;gap:6px;align-items:flex-start;background:#f9fafb;border:1px solid var(--line);padding:7px 9px;border-radius:5px;color:var(--ink)}.gatecheck input{min-width:0;margin-top:3px}.gatecheck span{display:grid}.gatecheck small{color:var(--muted)}.note-preview{white-space:pre-wrap;background:#f9fafb;border:1px solid var(--line);padding:10px;border-radius:5px;font:12px/1.45 Consolas,monospace}.card button{background:var(--blue);color:#fff;border:0;border-radius:5px;padding:8px 10px;cursor:pointer}"
    (assets / "site.css").write_text(css, encoding="utf-8")

    js = r"""const D=window.RESULTS_DATA;const $=(s)=>document.querySelector(s);const money=n=>'$'+(Math.abs(n)>=1e9?(n/1e9).toFixed(2)+'B':Math.abs(n)>=1e6?(n/1e6).toFixed(1)+'M':n.toLocaleString(undefined,{maximumFractionDigits:0}));const integer=n=>Math.round(n).toLocaleString();const pct=n=>(100*n).toFixed(1)+'%';function totals(year,tier){return D.totals.filter(x=>x.year==year&&x.tier==tier)}function trusted(country,year){return D.totals.find(x=>x.country==country&&x.year==year&&x.tier==='Trusted')}function renderOverview(){if(!$('#country-cards'))return;const y=String(D.metadata.macro_year),v=trusted('Vietnam',y);$('#country-cards').innerHTML=['Vietnam','India','Pakistan'].map(c=>{const x=trusted(c,y),all=D.totals.filter(z=>z.country===c&&z.year===y).reduce((a,z)=>a+z.value,0);return `<article class="card"><span class="status">${c} FY${y}</span><div class="kpi">${money(x.value)}</div><p>Trusted mapped value · ${integer(x.rows)} rows</p><p class="muted">Trusted share of all import value: ${pct(x.value/all)}</p></article>`}).join('');const rows=['India','Pakistan'].map(c=>{const x=trusted(c,y),m=D.macro[c],vm=D.macro.Vietnam;const popExpected=v.value*m.population/vm.population,gdpExpected=v.value*m.gdp/vm.gdp;return `<tr><td>${c}</td><td class="num">${money(x.value)}</td><td class="num">${(m.population/vm.population).toFixed(2)}×</td><td class="num">${money(popExpected)}</td><td class="num flag-${x.value/popExpected<.25?'low':x.value/popExpected<.6?'mid':'ok'}">${pct(x.value/popExpected)}</td><td class="num">${(m.gdp/vm.gdp).toFixed(2)}×</td><td class="num">${money(gdpExpected)}</td><td class="num flag-${x.value/gdpExpected<.25?'low':x.value/gdpExpected<.6?'mid':'ok'}">${pct(x.value/gdpExpected)}</td></tr>`}).join('');$('#benchmark-body').innerHTML=rows}
function renderComparison(){if(!$('#comparison-body'))return;const year=$('#year').value,tier=$('#tier').value,dim=$('#dimension').value,q=$('#search').value.toLowerCase();let rows=D.detail.filter(x=>x.year===year&&x.tier===tier&&x.dimension===dim&&x.name.toLowerCase().includes(q));const map={};for(const x of rows)(map[x.name]??={name:x.name,Vietnam:0,India:0,Pakistan:0,vr:0,ir:0,pr:0})[x.country]=x.value,(map[x.name][x.country[0].toLowerCase()+'r']=x.rows);rows=Object.values(map).sort((a,b)=>(b.Vietnam+b.India+b.Pakistan)-(a.Vietnam+a.India+a.Pakistan));const max=Math.max(1,...rows.map(x=>x.Vietnam+x.India+x.Pakistan));$('#row-count').textContent=`${integer(rows.length)} categories`;$('#comparison-body').innerHTML=rows.map(x=>`<tr><td>${x.name}</td><td class="num">${money(x.Vietnam)}</td><td class="num">${money(x.India)}</td><td class="num">${money(x.Pakistan)}</td><td class="num">${x.Vietnam? (x.India/x.Vietnam).toFixed(2)+'×':'—'}</td><td class="num">${x.Vietnam? (x.Pakistan/x.Vietnam).toFixed(2)+'×':'—'}</td><td><div class="bar"><i style="width:${100*(x.Vietnam+x.India+x.Pakistan)/max}%"></i></div></td></tr>`).join('')||'<tr><td colspan="7">No matching categories.</td></tr>'}function initComparison(){if(!$('#comparison-body'))return;['year','tier','dimension','search'].forEach(id=>$('#'+id).addEventListener(id==='search'?'input':'change',renderComparison));renderComparison()}
function renderSchemas(){if(!$('#schemas'))return;$('#schemas').innerHTML=D.schemas.map(w=>`<section class="schema"><h2>${w.file}</h2>${w.sheets.map(s=>`<details><summary>${s.name} · ${integer(Math.max(0,s.rows-1))} data rows · ${s.columns.length} columns</summary><div class="inside"><div class="columns">${s.columns.filter(Boolean).map(c=>`<code>${c}</code>`).join('')}</div></div></details>`).join('')}</section>`).join('')}
const simState={enabled:{}};const msum=xs=>xs.reduce((a,x)=>a.map((v,i)=>v+(x[i]||0)),[0,0,0]);function simFiles(){const s=$('#sim-scope').value;return s==='ALL'?[...new Set(D.totals.map(x=>x.file))]:[s]}function simRun(){const files=simFiles(),enabledMask=D.simulator.gates.reduce((m,g)=>m+(simState.enabled[g.key]?g.bit:0),0);let direct=[0,0,0],still=[0,0,0];const caught={};D.simulator.gates.forEach(g=>caught[g.key]=[0,0,0]);for(const x of D.simulator.groups.filter(x=>files.includes(x.file))){if(simState.enabled[x.gate])continue;const remaining=x.mask&enabledMask;if(!remaining)direct=msum([direct,x.m]);else{still=msum([still,x.m]);const g=D.simulator.gates.find(g=>(remaining&g.bit)!==0);caught[g.key]=msum([caught[g.key],x.m])}}const base=msum(D.totals.filter(x=>files.includes(x.file)&&x.tier==='Trusted').map(x=>[x.rows,x.value,x.volume]));const locked=msum(files.map(f=>(D.simulator.locked[f]||{m:[0,0,0]}).m));return{base,direct,still,caught,locked}}function renderSimulator(){if(!$('#sim-results'))return;const r=simRun();$('#sim-results').innerHTML=`<div class="grid"><article class="card"><span class="status">Simulated Trusted rows</span><div class="kpi">${integer(r.base[0]+r.direct[0])}</div><p>${r.direct[0]?'+':''}${integer(r.direct[0])} from the tested gate changes</p></article><article class="card"><span class="status">Simulated Trusted value</span><div class="kpi">${money(r.base[1]+r.direct[1])}</div><p>${r.direct[1]?'+':''}${money(r.direct[1])} from the tested gate changes</p></article><article class="card"><span class="status">Simulated Trusted volume</span><div class="kpi">${integer(r.base[2]+r.direct[2])}</div><p>${r.direct[2]?'+':''}${integer(r.direct[2])} from the tested gate changes</p></article></div><div class="notice"><b>Still held elsewhere:</b> ${integer(r.still[0])} rows / ${money(r.still[1])}. <b>Locked, non-toggleable:</b> ${integer(r.locked[0])} rows / ${money(r.locked[1])}.</div><div class="table-wrap"><table><thead><tr><th>Enabled gate likely to hold a released row</th><th class="num">Rows</th><th class="num">Value</th><th class="num">Volume</th></tr></thead><tbody>${D.simulator.gates.map(g=>`<tr><td>${g.label}</td><td class="num">${integer(r.caught[g.key][0])}</td><td class="num">${money(r.caught[g.key][1])}</td><td class="num">${integer(r.caught[g.key][2])}</td></tr>`).join('')}</tbody></table></div>`}function initSimulator(){if(!$('#sim-gates'))return;D.simulator.gates.forEach(g=>simState.enabled[g.key]=true);const files=[...new Set(D.totals.map(x=>x.file))];$('#sim-scope').innerHTML='<option value="ALL">All market-years</option>'+files.map(f=>`<option value="${f}">${f}</option>`).join('');$('#sim-gates').innerHTML=D.simulator.gates.map(g=>`<label class="gatecheck"><input type="checkbox" data-gate="${g.key}" checked><span><b>${g.label}</b><small>${g.stage}</small></span></label>`).join('');$('#sim-scope').addEventListener('change',renderSimulator);$('#sim-gates').addEventListener('change',e=>{if(e.target.dataset.gate){simState.enabled[e.target.dataset.gate]=e.target.checked;renderSimulator()}});renderSimulator()}document.addEventListener('DOMContentLoaded',()=>{renderOverview();initComparison();initSimulator();renderSchemas()});"""
    js += r"""
function simNote(){const r=simRun(),scope=$('#sim-scope').value==='ALL'?'all market-years':$('#sim-scope').value;const disabled=D.simulator.gates.filter(g=>!simState.enabled[g.key]).map(g=>g.label);return `Mapping gate what-if request\nScope: ${scope}\nGate(s) tested: ${disabled.length?disabled.join(', '):'none'}\nPotential direct release: ${integer(r.direct[0])} rows / ${money(r.direct[1])} / ${integer(r.direct[2])} volume\nLikely still held by another gate: ${integer(r.still[0])} rows / ${money(r.still[1])}\nLocked and not toggleable: ${integer(r.locked[0])} rows / ${money(r.locked[1])}\nBusiness rationale: [add here]\nRequested analyst action: inspect the all-row Excel examples, then approve or reject through adjudication.\n\nThis is a discussion simulation, not a production change.`}function refreshNote(){if($('#sim-note-preview'))$('#sim-note-preview').textContent=simNote()}document.addEventListener('DOMContentLoaded',()=>{if(!$('#sim-note-preview'))return;refreshNote();$('#sim-scope').addEventListener('change',refreshNote);$('#sim-gates').addEventListener('change',refreshNote);$('#copy-note').addEventListener('click',async()=>{await navigator.clipboard.writeText(simNote());$('#copy-note').textContent='Copied'});$('#download-note').addEventListener('click',()=>{const a=document.createElement('a');a.href=URL.createObjectURL(new Blob([simNote()],{type:'text/plain'}));a.download='mapping-gate-what-if-note.txt';a.click();URL.revokeObjectURL(a.href)})});
"""
    (assets / "site.js").write_text(js, encoding="utf-8")

    overview = """<div class="eyebrow">Current results at a glance</div><h1>Understand the mapped results before opening a workbook</h1><p class="muted">One governed entry point for aggregate market totals, category comparisons, output tracking, and field definitions.</p><div id="country-cards" class="grid"></div><h2>Vietnam benchmark sense-check</h2><p>FY2025 Trusted value is compared with what India and Pakistan would show if they scaled exactly with Vietnam by 2025 population or nominal GDP.</p><div class="notice"><b>Interpretation:</b> this is a reasonableness signal, not a market-size forecast. Trade capture, local manufacturing, coding practice, healthcare intensity, pricing, and mapping recall differ by country.</div><div class="table-wrap"><table><thead><tr><th>Market</th><th class="num">Actual Trusted</th><th class="num">Population vs VN</th><th class="num">Population-scaled</th><th class="num">Actual / expected</th><th class="num">GDP vs VN</th><th class="num">GDP-scaled</th><th class="num">Actual / expected</th></tr></thead><tbody id="benchmark-body"></tbody></table></div><h2>Where to go</h2><div class="grid"><article class="card"><h3>Weekend scorecard</h3><p>See the measured recall movement, the precision estimate, examples, and why the change happened.</p><a href="quality.html">Open scorecard →</a></article><article class="card"><h3>What-if playground</h3><p>Turn aggregate gates on or off and see how Trusted totals would move without changing production.</p><a href="simulator.html">Open playground →</a></article><article class="card"><h3>Compare results</h3><p>Slice exact Trusted, Review, or Excluded value by family, manufacturer, product, or business hierarchy.</p><a href="comparison.html">Open comparison →</a></article><article class="card"><h3>Outputs & tracking</h3><p>Choose between all-row raw outputs and aggregate-only dashboards.</p><a href="outputs.html">Open directory →</a></article><article class="card"><h3>Data schemas</h3><p>See every sheet and column without loading the row-level files.</p><a href="schemas.html">Open schemas →</a></article></div>"""
    (out / "index.html").write_text(shell("Results navigation", "Overview", overview), encoding="utf-8")

    recall = scorecard.get("realized_recall", {})
    net = recall.get("net_trusted", {})
    newly = recall.get("newly_trusted", {})
    lost = recall.get("lost_trusted", {})
    population = scorecard.get("population", {})
    baseline_trusted = next((x["baseline"] for x in scorecard.get("tier_totals", [])
                             if x["file"] == "OVERALL" and x["tier"] == "Trusted"), {})
    row_gain_pct = (net.get("rows", 0) / baseline_trusted.get("rows", 1)) * 100
    value_gain_pct = (net.get("value_usd", 0) / baseline_trusted.get("value_usd", 1)) * 100
    volume_gain_pct = (net.get("volume", 0) / baseline_trusted.get("volume", 1)) * 100
    precision_rows = review.get("aggregate", {}).get("precision_scores", [])
    trusted_precision = next((x for x in precision_rows
                              if x.get("sample_type") == "Deterministic stratified random"
                              and x.get("scope") == "Trusted"), {})
    top_rows = "".join(
        f"<tr><td>{esc(x.get('proposed_player') or 'Unspecified')} — {esc(x.get('proposed_family') or x.get('alias_term'))}</td>"
        f"<td class='num'>{fmt_number(x.get('cluster_rows', 0))}</td>"
        f"<td class='num'>{fmt_money(x.get('cluster_value_usd', 0))}</td>"
        f"<td>{'Family alias' if x.get('proposal_type') == 'family_alias' else 'Scope whitelist'} passed two independent reviews</td></tr>"
        for x in review.get("top_approved", [])[:8]
    ) or "<tr><td colspan='4'>No approved proposal aggregates are available.</td></tr>"
    quality = f"""<div class="eyebrow">Weekend change, in plain English</div><h1>Recall and precision scorecard</h1><p class="muted">The comparison uses {fmt_number(population.get('common_rows', 0))} exactly matched source rows from last week's and the current audits. Precision is an LLM estimate until business reviewers label the sample.</p><h2>What the recall work recovered</h2><div class="grid"><article class="card"><span class="status">Gross row recovery</span><div class="kpi">+{fmt_number(newly.get('rows', 0))}</div><p>previously blocked rows moved into Trusted</p></article><article class="card"><span class="status">Gross value recovery</span><div class="kpi">+{fmt_money(newly.get('value_usd', 0))}</div><p>previously blocked value moved into Trusted</p></article><article class="card"><span class="status">Gross volume recovery</span><div class="kpi">+{fmt_number(newly.get('volume', 0))}</div><p>previously blocked units moved into Trusted</p></article></div><h2>Net Trusted change after safeguards</h2><div class="grid"><article class="card"><span class="status">Rows</span><div class="kpi">{fmt_number(net.get('rows', 0))}</div><p>net change in Trusted rows ({row_gain_pct:+.2f}%)</p></article><article class="card"><span class="status">Value</span><div class="kpi">{fmt_money(net.get('value_usd', 0))}</div><p>net change in Trusted value ({value_gain_pct:+.2f}%)</p></article><article class="card"><span class="status">Volume</span><div class="kpi">{fmt_number(net.get('volume', 0))}</div><p>net change in Trusted units ({volume_gain_pct:+.2f}%)</p></article></div><div class="notice"><b>What this means:</b> the recall work genuinely recovered {fmt_number(newly.get('rows', 0))} rows, but separate quality safeguards held back {fmt_number(lost.get('rows', 0))} previously Trusted rows pending safer validation. Therefore this weekend improved gross recall and defensibility, but it did not produce a net increase in Trusted coverage.</div><h2>Precision estimate</h2><div class="grid"><article class="card"><h3>Trusted row precision</h3><div class="kpi">{trusted_precision.get('estimated_mapping_precision_rows', 0) * 100:.2f}%</div><p>Estimated correctly mapped Trusted rows</p></article><article class="card"><h3>Value-weighted precision</h3><div class="kpi">{trusted_precision.get('estimated_mapping_precision_value', 0) * 100:.2f}%</div><p>Estimated share of Trusted value mapped correctly</p></article><article class="card"><h3>Volume-weighted precision</h3><div class="kpi">{trusted_precision.get('estimated_mapping_precision_volume', 0) * 100:.2f}%</div><p>Estimated share of Trusted volume mapped correctly</p></article></div><div class="notice"><b>Caution:</b> these precision figures come from only {fmt_number(trusted_precision.get('sample_rows', 0))} randomly sampled Trusted rows reviewed by LLMs. They are decision-support estimates, not human-verified ground truth.</div><h2>Why “mAP” is not shown</h2><p>Mean Average Precision is designed for ranked search or object-detection results. This workflow assigns each row to Trusted, Review, or Excluded; it does not return a ranked list. Reporting mAP would therefore be misleading. The closest useful business measures are the exact Trusted gains above plus row-, value-, and volume-weighted precision.</p><h2>What changed</h2><div class="grid"><article class="card"><h3>Strict LLM adjudication</h3><div class="kpi">{fmt_number(review.get('totals', {}).get('approved', 0))}</div><p>of {fmt_number(review.get('totals', {}).get('reviewed', 0))} proposals passed dual-review consensus.</p></article><article class="card"><h3>Governed opportunity</h3><div class="kpi">{fmt_money(review.get('totals', {}).get('approved_value', 0))}</div><p>Candidate cluster value approved for governed reference updates; realized gain is reported separately above.</p></article><article class="card"><h3>Processing design</h3><div class="kpi">SQLite</div><p>All review, reconciliation, and audit calculations use a database. Excel is reserved for final all-row handoff files.</p></article></div><h2>Aggregate examples that passed the stricter review</h2><div class="table-wrap"><table><thead><tr><th>Manufacturer — family / term</th><th class="num">Cluster rows</th><th class="num">Candidate value</th><th>Why it progressed</th></tr></thead><tbody>{top_rows}</tbody></table></div><p class="muted">These are proposal-cluster aggregates, not row-level dashboard data. See Raw outputs for the complete audit trail.</p>"""
    (out / "quality.html").write_text(shell("Weekend recall and precision", "Weekend scorecard", quality), encoding="utf-8")

    simulator_body = f"""<div class="eyebrow">Aggregate discussion tool</div><h1>What-if gate playground</h1><p class="muted">Uncheck a gate to estimate what would happen if that gate released its currently blocked rows. Results update instantly and remain grouped statistics only.</p><div class="notice"><b>Simulation only — no production change.</b> Secondary gates may still hold a released row; recovery dynamics and downstream remapping are not modelled. Any insight must go through business adjudication and a governed rerun.</div><div class="toolbar"><label>Market-year<select id="sim-scope"></select></label><div id="sim-gates" class="gates"></div></div><div id="sim-results"></div><h2>Turn the scenario into an adjudication request</h2><p>Copy or download the aggregate scenario note, add the business rationale, then inspect concrete rows in <a href="{esc(raw_link_prefix)}/LLM_Review_Raw_Output.xlsx">LLM Review Raw Output</a> or <a href="{esc(raw_link_prefix)}/Recall_Recovery_Proposals_LLM_Reviewed.xlsx">Recall Recovery Proposals</a>. Only approved changes enter the governed reference lists and production rerun.</p><div class="grid"><article class="card"><h3>Suggested request note</h3><pre id="sim-note-preview" class="note-preview"></pre><button id="copy-note" type="button">Copy note</button> <button id="download-note" type="button">Download note</button></article><article class="card"><h3>Business review checklist</h3><p>Check that the family and manufacturer make sense, the mapped product is in scope, and the same change works across several examples—not just one unusual row.</p></article><article class="card"><h3>Governance</h3><p>The playground never edits production. Analysts decide in Excel; the approved list is ingested into reference tables and the full pipeline is rerun and verified.</p></article></div>"""
    (out / "simulator.html").write_text(shell("What-if gate playground", "What-if playground", simulator_body), encoding="utf-8")

    comparison = f"""<div class="eyebrow">Exact cross-market comparison</div><h1>Compare totals, families, and manufacturers</h1><p class="muted">All values reconcile to {esc(audit_label)}. Blank mapped dimensions are shown as &lt;Unmapped&gt;; genuine Unspecified labels remain distinct.</p><div class="toolbar"><label>Fiscal year<select id="year"><option>2025</option><option>2024</option></select></label><label>Routing tier<select id="tier"><option>Trusted</option><option>Review</option><option>Excluded</option></select></label><label>Breakdown<select id="dimension"><option value="family">Family</option><option value="manufacturer">Manufacturer</option><option value="product">Product</option><option value="segment">Segment</option><option value="sub_segment">Sub-segment</option></select></label><label>Find category<input id="search" type="search" placeholder="Type to filter"></label><span id="row-count" class="status"></span></div><div class="table-wrap"><table><thead><tr><th>Category</th><th class="num">Vietnam</th><th class="num">India</th><th class="num">Pakistan</th><th class="num">India / VN</th><th class="num">Pakistan / VN</th><th>Combined scale</th></tr></thead><tbody id="comparison-body"></tbody></table></div><div class="notice"><b>How to use:</b> start with Trusted, then switch to Review to see where an apparent market gap may be a mapping-recall gap rather than a real commercial difference.</div>"""
    (out / "comparison.html").write_text(shell("Compare results", "Compare results", comparison), encoding="utf-8")

    raw_links = "".join(
        f'<a href="{esc(raw_link_prefix)}/mapped_results/{esc(s["file"])}">{esc(s["file"].replace("_ML_Map_Mapped.xlsx", ""))}</a>'
        for s in schemas
    )
    outputs = f"""<div class="eyebrow">Current governed outputs</div><h1>Two deliberately separate deliverables</h1><p class="muted">Use the dashboard to understand the result; open Raw outputs only when you need the complete row-level evidence.</p><div class="grid links"><article class="card"><span class="status">1 · Raw outputs</span><h3>All rows · Excel final output</h3><p>Six final Excel workbooks contain every source row, split across Excel sheets where the Excel row limit requires it.</p>{raw_links}<a href="{esc(raw_link_prefix)}/LLM_Review_Raw_Output.xlsx">Complete LLM review output</a><a href="{esc(raw_link_prefix)}/Recall_Recovery_Proposals_LLM_Reviewed.xlsx">Reviewed recall proposals</a><a href="{esc(raw_link_prefix)}/Prediction_Funnel_and_Review.xlsx">Business review workbook</a></article><article class="card"><span class="status">2 · Dashboard</span><h3>Aggregate statistics only · HTML</h3><p>This navigation site contains totals and grouped statistics by country, year, family, manufacturer, product, segment, sub-segment, and gate-mask scenario. It contains no row-level records.</p><a href="index.html">Aggregate overview</a><a href="quality.html">Recall & precision scorecard</a><a href="simulator.html">What-if playground</a><a href="comparison.html">Family / manufacturer comparison</a><a href="schemas.html">Schema dictionary</a></article><article class="card"><h3>Stable processing</h3><p>Back-end review, reconciliation, and audit calculations use SQLite to avoid Excel stability and row-limit problems. SQLite is the internal processing backend and authority, not a final business output.</p></article><article class="card"><h3>Tracking & lineage</h3><p>What changed, current output bounds, and source lineage.</p><a href="../../5. Documentation/DATA_UPDATES_LOG.md">Data updates log</a><a href="../../5. Documentation/OUTPUT_TRACKER.md">Output tracker</a><a href="../../5. Documentation/DATA_LINEAGE.md">Data lineage</a></article></div>"""
    (out / "outputs.html").write_text(shell("Outputs and tracking", "Outputs & tracking", outputs), encoding="utf-8")

    schemas_body = """<div class="eyebrow">Workbook dictionary</div><h1>Current mapped-output schemas</h1><p class="muted">Expand a worksheet to see its exact column names and aggregate row count. No data rows are embedded in this page.</p><div class="notice"><b>Core distinction:</b> RawData (and RawData_Part_2, etc.) together contain all source rows; Trusted_Dashboard contains only rows passing release gates; Review_Queue holds unresolved rows and must not be added to market totals.</div><div id="schemas"></div>"""
    (out / "schemas.html").write_text(shell("Data schemas", "Data schemas", schemas_body), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", default=str(DEFAULT_DB))
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    parser.add_argument("--workbook-dir", type=Path, default=DEFAULT_WB_DIR)
    parser.add_argument("--scorecard", type=Path,
                        help="Aggregate JSON from compare_prediction_audits.py")
    parser.add_argument("--review-db", type=Path, default=DEFAULT_REVIEW_DB)
    parser.add_argument("--review-aggregate", type=Path,
                        help="Aggregate-only LLM review JSON")
    parser.add_argument("--audit-label", default="current governed audit")
    parser.add_argument("--raw-link-prefix", default="../../raw_outputs",
                        help="Relative URL from the site to final Excel raw outputs")
    args = parser.parse_args()
    conn = sqlite3.connect(f"file:{Path(args.db).resolve().as_posix()}?mode=ro", uri=True)
    totals, detail = aggregate(conn)
    file_ids = sorted({x["file"] for x in totals})
    simulator = _build_simulator(conn.cursor(), file_ids)
    for group in simulator["groups"]:
        group.pop("examples", None)
    conn.close()
    review = load_review_summary(args.review_db)
    if args.review_aggregate and args.review_aggregate.exists():
        review["aggregate"] = json.loads(args.review_aggregate.read_text(encoding="utf-8"))
    write_site(
        Path(args.out), totals, detail, workbook_schemas(args.workbook_dir),
        load_scorecard(args.scorecard), review, args.audit_label, args.raw_link_prefix,
        simulator,
    )
    print(f"Built results navigation: {Path(args.out).resolve()}")
    print(f"Totals: {len(totals)}; comparison groups: {len(detail)}")


if __name__ == "__main__":
    main()
