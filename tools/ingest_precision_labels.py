#!/usr/bin/env python3
"""Validate and ingest business precision labels into the audit authority.

This tool reads the governed ``Review Samples`` worksheet and updates only the
reviewer-owned label fields in ``review_label``. It never edits the workbook,
production outputs, routing, reference lists, or adjudication fields.

Examples:
    python tools/ingest_precision_labels.py --check --workbook <shared workbook>
    python tools/ingest_precision_labels.py --workbook <shared workbook>
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import sqlite3
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
DEFAULT_RUN = "20260710_recall_audit_v2"
DEFAULT_DB = REPO / "outputs" / DEFAULT_RUN / "prediction_audit.sqlite"
DEFAULT_WORKBOOK = REPO / "outputs" / DEFAULT_RUN / "Prediction_Funnel_and_Review.xlsx"
SHEET = "Review Samples"
HEADER_ROW = 4
TOOL_VERSION = "1.0"

FIELDS = (
    "surgical_relevance", "mapping_correctness", "corrected_manufacturer",
    "corrected_family", "corrected_product", "corrected_segment",
    "corrected_sub_segment", "reviewer_rationale", "reviewer", "reviewed_at",
)
HEADER_TO_KEY = {
    "Output ID": "output_file_id",
    "Source row ID": "source_row_id",
    "Surgical Relevance": "surgical_relevance",
    "Mapping Correctness": "mapping_correctness",
    "Corrected Manufacturer": "corrected_manufacturer",
    "Corrected Family": "corrected_family",
    "Corrected Product": "corrected_product",
    "Corrected Segment": "corrected_segment",
    "Corrected Sub Segment": "corrected_sub_segment",
    "Reviewer Rationale": "reviewer_rationale",
    "Reviewer": "reviewer",
    "Reviewed at": "reviewed_at",
}
RELEVANCE_VALUES = {"Surgical", "Not surgical", "Uncertain"}
MAPPING_VALUES = {"Correct", "Incorrect", "Uncertain"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(4 * 1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat(timespec="seconds") if isinstance(value, dt.datetime) else value.isoformat()
    text = str(value).strip()
    return text or None


def clean_id(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value) or ""


def read_labels(path: Path) -> list[dict[str, str | None]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            raise ValueError(f"Workbook has no {SHEET!r} sheet.")
        ws = wb[SHEET]
        headers = [clean(c.value) for c in ws[HEADER_ROW]]
        missing = sorted(set(HEADER_TO_KEY) - set(headers))
        if missing:
            raise ValueError(f"Review sheet is missing required columns: {', '.join(missing)}")
        positions = {HEADER_TO_KEY[h]: i for i, h in enumerate(headers) if h in HEADER_TO_KEY}
        rows: list[dict[str, str | None]] = []
        for cells in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            fid = clean_id(cells[positions["output_file_id"]])
            source_id = clean_id(cells[positions["source_row_id"]])
            if not fid and not source_id:
                continue
            row: dict[str, str | None] = {
                "output_file_id": fid,
                "source_row_id": source_id,
            }
            for field in FIELDS:
                row[field] = clean(cells[positions[field]])
            rows.append(row)
        return rows
    finally:
        wb.close()


def validate(rows: list[dict[str, str | None]], authority_keys: set[tuple[str, str]]) -> list[str]:
    errors: list[str] = []
    keys = [(str(r["output_file_id"]), str(r["source_row_id"])) for r in rows]
    if len(keys) != len(set(keys)):
        errors.append("The workbook contains duplicate Output ID × Source row ID keys.")
    if set(keys) != authority_keys:
        missing = authority_keys - set(keys)
        extra = set(keys) - authority_keys
        errors.append(f"Workbook/authority identities differ: {len(missing)} missing, {len(extra)} unexpected.")
    for row in rows:
        key = f"{row['output_file_id']} / row {row['source_row_id']}"
        rel = row["surgical_relevance"]
        mapping = row["mapping_correctness"]
        if rel and rel not in RELEVANCE_VALUES:
            errors.append(f"{key}: invalid Surgical Relevance {rel!r}.")
        if mapping and mapping not in MAPPING_VALUES:
            errors.append(f"{key}: invalid Mapping Correctness {mapping!r}.")
        activity = any(row[f] for f in FIELDS)
        if not activity:
            continue
        if not rel:
            errors.append(f"{key}: Surgical Relevance is required once a row is reviewed.")
        if rel == "Surgical" and not mapping:
            errors.append(f"{key}: Mapping Correctness is required for a Surgical row.")
        if mapping and not rel:
            errors.append(f"{key}: Mapping Correctness cannot be entered without Surgical Relevance.")
        if not row["reviewer"] or not row["reviewed_at"]:
            errors.append(f"{key}: Reviewer and Reviewed at are required for traceability.")
        if mapping == "Incorrect":
            corrected = any(row[f] for f in (
                "corrected_manufacturer", "corrected_family", "corrected_product",
                "corrected_segment", "corrected_sub_segment",
            ))
            if not corrected:
                errors.append(f"{key}: an Incorrect mapping needs at least one corrected field.")
            if not row["reviewer_rationale"]:
                errors.append(f"{key}: an Incorrect mapping needs a reviewer rationale.")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db", default=str(DEFAULT_DB))
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--check", action="store_true", help="Validate only; do not update SQLite.")
    args = parser.parse_args()
    db = Path(args.db).resolve()
    workbook = Path(args.workbook).resolve()
    if not db.exists() or not workbook.exists():
        print(f"ERROR: missing {'database' if not db.exists() else 'workbook'}.", file=sys.stderr)
        return 2

    rows = read_labels(workbook)
    uri = f"file:{db.as_posix()}?mode=ro" if args.check else str(db)
    con = sqlite3.connect(uri, uri=args.check)
    try:
        run_rows = con.execute("SELECT run_id FROM run").fetchall()
        if len(run_rows) != 1:
            raise RuntimeError(f"Expected exactly one audit run, found {len(run_rows)}.")
        run_id = str(run_rows[0][0])
        authority_keys = {
            (str(fid), str(source_id))
            for fid, source_id in con.execute("SELECT output_file_id,source_row_id FROM review_label")
        }
        errors = validate(rows, authority_keys)
        labelled = [r for r in rows if any(r[f] for f in FIELDS)]
        print(f"Workbook: {workbook}")
        print(f"Audit run: {run_id}")
        print(f"Sample identities: {len(rows)} workbook / {len(authority_keys)} authority")
        print(f"Rows with reviewer content: {len(labelled)}")
        if errors:
            print(f"VALIDATION FAILED ({len(errors)} issue(s)):", file=sys.stderr)
            for error in errors[:50]:
                print(f"  - {error}", file=sys.stderr)
            if len(errors) > 50:
                print(f"  - ... {len(errors) - 50} more", file=sys.stderr)
            return 1
        if args.check:
            print("VALIDATION PASSED — check-only; SQLite was not changed.")
            return 0

        digest = sha256_file(workbook)
        now = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()
        with con:
            con.execute(
                """CREATE TABLE IF NOT EXISTS review_label_ingest_batch (
                       batch_id INTEGER PRIMARY KEY,
                       ingested_at TEXT NOT NULL,
                       workbook_path TEXT NOT NULL,
                       workbook_sha256 TEXT NOT NULL,
                       db_run_id TEXT NOT NULL,
                       workbook_rows INTEGER NOT NULL,
                       labelled_rows INTEGER NOT NULL,
                       tool_version TEXT NOT NULL,
                       UNIQUE(db_run_id, workbook_sha256)
                   )"""
            )
            prior = con.execute(
                "SELECT batch_id FROM review_label_ingest_batch WHERE db_run_id=? AND workbook_sha256=?",
                (run_id, digest),
            ).fetchone()
            if prior:
                print(f"NO-OP — this exact workbook was already ingested as batch {prior[0]}.")
                return 0
            for row in labelled:
                con.execute(
                    """UPDATE review_label SET
                           surgical_relevance=?, mapping_correctness=?,
                           corrected_manufacturer=?, corrected_family=?, corrected_product=?,
                           corrected_segment=?, corrected_sub_segment=?, reviewer_rationale=?,
                           reviewer=?, reviewed_at=?
                         WHERE output_file_id=? AND source_row_id=?""",
                    tuple(row[f] for f in FIELDS)
                    + (row["output_file_id"], row["source_row_id"]),
                )
            con.execute(
                """INSERT INTO review_label_ingest_batch
                   (ingested_at,workbook_path,workbook_sha256,db_run_id,
                    workbook_rows,labelled_rows,tool_version)
                   VALUES (?,?,?,?,?,?,?)""",
                (now, str(workbook), digest, run_id, len(rows), len(labelled), TOOL_VERSION),
            )
        print(f"INGESTED — {len(labelled)} reviewed row(s); production routing unchanged.")
        return 0
    finally:
        con.close()


if __name__ == "__main__":
    raise SystemExit(main())
