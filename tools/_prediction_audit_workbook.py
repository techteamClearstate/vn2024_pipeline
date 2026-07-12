"""Pure-Python (openpyxl) builder for `Prediction_Funnel_and_Review.xlsx`.

Replaces the bundled node/Excel artifact-tool (which aborts with a V8 error in some
environments) with a dependency-light openpyxl implementation that produces the SAME
seven governed sheets in the SAME order that `verify_prediction_audit.py` asserts:

    Read Me · Funnel · Removal Cube · Review Samples · Recall Risks ·
    Reconciliation QC · Source Lineage

It consumes the exact `payload` dict produced by
`build_prediction_audit_reports.build_payload()`. Review-only artifact; no authority is
mutated here.
"""
from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAX_WORKBOOK_BYTES = 100_000_000

INK = "0E0E0E"
BLUE = "0047AB"
BLUE_LIGHT = "EAF1FB"
GRAY_LIGHT = "F3F4F6"
WHITE = "FFFFFF"
GREEN = "D9EAD3"
AMBER = "FFF2CC"
RED = "F4CCCC"
LINE = "D8DEE8"

_THIN = Side(style="thin", color=LINE)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _norm(v):
    if isinstance(v, bool):
        return 1 if v else 0
    return v


def _title(ws, title, note, ncols):
    lc = get_column_letter(ncols)
    ws.merge_cells(f"A1:{lc}1")
    ws.merge_cells(f"A2:{lc}2")
    a1 = ws["A1"]
    a1.value = title
    a1.fill = _fill(INK)
    a1.font = Font(bold=True, color=WHITE, size=18)
    a1.alignment = Alignment(vertical="center")
    a2 = ws["A2"]
    a2.value = note
    a2.fill = _fill(BLUE_LIGHT)
    a2.font = Font(italic=True, color=INK)
    a2.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 34


def _widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def _header(ws, row, labels):
    fill = _fill(BLUE)
    font = Font(bold=True, color=WHITE)
    for j, label in enumerate(labels):
        c = ws.cell(row=row, column=j + 1, value=label)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = _BORDER
    ws.row_dimensions[row].height = 30


def _data_sheet(ws, title, note, columns, rows, widths,
                number_formats=None, dropdowns=None, cond_col=None):
    number_formats = number_formats or {}
    ncols = len(columns)
    lc = get_column_letter(ncols)
    _title(ws, title, note, ncols)
    _header(ws, 4, [c["label"] for c in columns])
    keys = [c["key"] for c in columns]
    first = 5
    for r, row in enumerate(rows):
        er = first + r
        for j, k in enumerate(keys):
            cell = ws.cell(row=er, column=j + 1, value=_norm(row.get(k)))
            fmt = number_formats.get(j)
            if fmt:
                cell.number_format = fmt
    last = first + len(rows) - 1
    if rows:
        ws.auto_filter.ref = f"A4:{lc}{last}"
    ws.freeze_panes = "A5"
    _widths(ws, widths)
    if dropdowns and rows:
        for col_idx, values in dropdowns.items():
            if col_idx < 0:
                continue
            letter = get_column_letter(col_idx + 1)
            dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{letter}{first}:{letter}{last}")
    if cond_col is not None and cond_col >= 0 and rows:
        letter = get_column_letter(cond_col + 1)
        rng = f"{letter}{first}:{letter}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"FAIL"'],
                                      fill=_fill(RED), font=Font(bold=True, color=INK)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"WARN"'],
                                      fill=_fill(AMBER), font=Font(bold=True, color=INK)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PASS"'],
                                      fill=_fill(GREEN), font=Font(bold=True, color=INK)))
    return last


# ---- column widths / number formats mirrored from the retired mjs builder ----
FUNNEL_WIDTHS = [14, 18, 12, 24, 10, 28, 12, 16, 22, 14, 16, 14, 12, 12, 12, 12, 14, 14, 16, 14, 14, 16, 14, 14]
FUNNEL_FMT = {9: "#,##0", 10: "$#,##0.00", 11: "#,##0.000000", 12: "#,##0", 13: "#,##0", 14: "#,##0",
              15: "#,##0", 16: "$#,##0.00", 17: "#,##0", 18: "$#,##0.00", 19: "#,##0.000000", 20: "#,##0",
              21: "$#,##0.00", 22: "#,##0.000000", 23: "0.00%"}
REMOVAL_WIDTHS = [14, 18, 12, 10, 24, 28, 12, 14, 28, 28, 34, 22, 22, 22, 14, 16, 14, 12, 12, 12, 12, 14, 14, 16, 14, 14, 16, 14, 14]
REMOVAL_FMT = {14: "#,##0", 15: "$#,##0.00", 16: "#,##0.000000", 17: "#,##0", 18: "#,##0", 19: "#,##0",
               20: "#,##0", 21: "$#,##0.00", 22: "#,##0", 23: "$#,##0.00", 24: "#,##0.000000", 25: "#,##0",
               26: "$#,##0.00", 27: "#,##0.000000", 28: "0.00%"}
QC_WIDTHS = [14, 22, 34, 22, 22, 14, 14, 12, 42, 22]
QC_FMT = {5: "$#,##0.000000", 6: "#,##0.000000"}


def _build_readme(ws, payload):
    meta = payload["meta"]
    _title(ws, "Prediction funnel audit — review-only authority",
           f"Run {meta['run_id']} | Registry {meta['registry_version']} | SQLite authority: {meta['sqlite_path']}", 8)
    r = 4
    for section in payload["read_me_sections"]:
        ws.merge_cells(f"A{r}:H{r}")
        c = ws.cell(row=r, column=1, value=section["title"])
        c.fill = _fill(BLUE)
        c.font = Font(bold=True, color=WHITE)
        r += 1
        for item in section["rows"]:
            ws.merge_cells(f"B{r}:H{r}")
            a = ws.cell(row=r, column=1, value=item[0])
            a.fill = _fill(GRAY_LIGHT)
            a.font = Font(bold=True, color=INK)
            a.alignment = Alignment(vertical="top", wrap_text=True)
            b = ws.cell(row=r, column=2, value=item[1])
            b.alignment = Alignment(vertical="top", wrap_text=True)
            r += 1
        r += 1
    ws.freeze_panes = "A3"
    _widths(ws, [24, 18, 18, 18, 18, 18, 18, 18])


def _build_recall(ws, payload):
    _title(ws, "Complete recall-risk inventory summary",
           "Counts are complete SQLite inventories, not samples. The workbook carries the aggregate plus "
           "high-value evidence; query recall_risk_inventory for every row-level record.", 9)
    _header(ws, 4, ["Output", "Label", "Risk type", "Current tier", "Transactions",
                    "Value USD", "Volume", "Weighted ASP", "Inventory scope"])
    r = 5
    for row in payload["recall_summary"]:
        vol = row.get("volume")
        val = row.get("value_usd")
        asp = (val / vol) if vol else None
        vals = [row.get("output_file_id"), row.get("output_label"), row.get("risk_type"),
                row.get("current_output_tier"), row.get("transaction_count"), val, vol, asp,
                "Complete SQLite inventory"]
        for j, v in enumerate(vals):
            cell = ws.cell(row=r, column=j + 1, value=_norm(v))
            if j == 4:
                cell.number_format = "#,##0"
            elif j in (5, 7):
                cell.number_format = "$#,##0.00"
            elif j == 6:
                cell.number_format = "#,##0.000000"
        r += 1
    summary_last = 4 + len(payload["recall_summary"])
    ev_header = summary_last + 3
    ev_cols = payload["recall_evidence_columns"]
    lc = get_column_letter(len(ev_cols))
    ws.merge_cells(f"A{ev_header}:I{ev_header}")
    c = ws.cell(row=ev_header, column=1,
                value="High-value evidence (bounded workbook projection; complete detail remains in SQLite)")
    c.fill = _fill(INK)
    c.font = Font(bold=True, color=WHITE)
    _header(ws, ev_header + 1, [c["label"] for c in ev_cols])
    keys = [c["key"] for c in ev_cols]
    r = ev_header + 2
    for row in payload["recall_evidence"]:
        for j, k in enumerate(keys):
            cell = ws.cell(row=r, column=j + 1, value=_norm(row.get(k)))
            if k == "value_usd":
                cell.number_format = "$#,##0.00"
        r += 1
    ws.freeze_panes = "A5"
    _widths(ws, [14, 18, 26, 14, 14, 16, 14, 16, 34])


def _build_lineage(ws, payload):
    _title(ws, "Source lineage and immutable baselines",
           "Absolute paths, SHA-256 hashes, row counts, and numeric-status totals identify the governed "
           "inputs used by this run.", 17)
    _header(ws, 4, ["Output", "Label", "Country", "FY", "Source path", "Complete-source path", "Format",
                    "Ingestion mode", "Completeness basis", "Expected rows", "Observed rows", "SHA-256",
                    "Bytes", "Value USD", "Volume", "Missing/invalid value", "Missing/invalid volume"])
    r = 5
    for row in payload["source_lineage"]:
        vals = [row.get("output_file_id"), row.get("output_label"), row.get("country"), row.get("fiscal_year"),
                row.get("source_path"), row.get("complete_source_path"), row.get("source_format"),
                row.get("ingestion_mode"), row.get("completeness_basis"), row.get("expected_rows"),
                row.get("observed_rows"), row.get("source_sha256"), row.get("source_bytes"),
                row.get("value_usd"), row.get("volume"),
                f"{row.get('missing_value_count')}/{row.get('invalid_value_count')}",
                f"{row.get('missing_volume_count')}/{row.get('invalid_volume_count')}"]
        for j, v in enumerate(vals):
            cell = ws.cell(row=r, column=j + 1, value=_norm(v))
            if j in (9, 10, 12):
                cell.number_format = "#,##0"
            elif j == 13:
                cell.number_format = "$#,##0.00"
            elif j == 14:
                cell.number_format = "#,##0.000000"
        r += 1
    source_last = 4 + len(payload["source_lineage"])
    man_header = source_last + 3
    ws.merge_cells(f"A{man_header}:Q{man_header}")
    c = ws.cell(row=man_header, column=1, value="Baseline manifest")
    c.fill = _fill(INK)
    c.font = Font(bold=True, color=WHITE)
    _header(ws, man_header + 1, ["Artifact type", "Path", "SHA-256", "Bytes", "Transactions",
                                 "Value USD", "Volume", "Run ID"])
    r = man_header + 2
    for row in payload["baseline_manifest"]:
        vals = [row.get("artifact_type"), row.get("path"), row.get("sha256"), row.get("bytes"),
                row.get("transaction_count"), row.get("value_usd"), row.get("volume"), row.get("run_id")]
        for j, v in enumerate(vals):
            cell = ws.cell(row=r, column=j + 1, value=_norm(v))
            if j == 3 or j == 4:
                cell.number_format = "#,##0"
            elif j == 5:
                cell.number_format = "$#,##0.00"
            elif j == 6:
                cell.number_format = "#,##0.000000"
        r += 1
    ws.freeze_panes = "A5"
    _widths(ws, [14, 18, 16, 10, 40, 40, 12, 34, 40, 14, 14, 34, 14, 16, 14, 18, 18])


def _review_widths(columns):
    wide = {"evidence", "shadow_recommendation", "detailed_product", "reviewer_rationale"}
    mid = {"source_row_id", "sample_stratum", "target_category", "primary_reason"}
    out = []
    for c in columns:
        if c["key"] in wide:
            out.append(34)
        elif c["key"] in mid:
            out.append(22)
        else:
            out.append(15)
    return out


def _index_of(columns, key):
    for i, c in enumerate(columns):
        if c["key"] == key:
            return i
    return -1


def build_workbook_xlsx(payload: dict, output_path: Path) -> Path:
    """Build the seven-sheet governed workbook with openpyxl and atomically replace."""
    wb = Workbook()
    readme = wb.active
    readme.title = "Read Me"
    readme.sheet_view.showGridLines = False
    _build_readme(readme, payload)

    _data_sheet(wb.create_sheet("Funnel"),
                "Candidate and terminal funnel",
                "Candidate states are not final routing. S13 is terminal routing; S14 is presentation only. "
                "Overall duplicates each physical row exactly once.",
                payload["funnel_columns"], payload["funnel"], FUNNEL_WIDTHS, FUNNEL_FMT)

    _data_sheet(wb.create_sheet("Removal Cube"),
                "Removal and risk cube",
                "Primary reasons are additive. Secondary reasons are nonadditive diagnostics and must not be "
                "summed into removals. <Unmapped> is distinct from a genuine Unspecified label.",
                payload["removal_columns"], payload["removal_cube"], REMOVAL_WIDTHS, REMOVAL_FMT)

    rcols = payload["review_columns"]
    rfmt = {}
    for k, code in (("inclusion_probability", "0.000000"), ("sample_weight", "0.000"),
                    ("value_usd", "$#,##0.00"), ("volume", "#,##0.000000")):
        idx = _index_of(rcols, k)
        if idx >= 0:
            rfmt[idx] = code
    dropdowns = {}
    si = _index_of(rcols, "surgical_relevance")
    mi = _index_of(rcols, "mapping_correctness")
    if si >= 0:
        dropdowns[si] = ["Surgical", "Not surgical", "Uncertain"]
    if mi >= 0:
        dropdowns[mi] = ["Correct", "Incorrect", "Uncertain"]
    _data_sheet(wb.create_sheet("Review Samples"),
                "Reviewer-ready deterministic sample",
                "Exactly 25 rows per output: 12 purposeful targets plus 13 deterministic stratified random "
                "rows. Purposeful rows do not support weighted population estimates. Recommendations are "
                "shadow-only.",
                rcols, payload["review_samples"], _review_widths(rcols), rfmt, dropdowns=dropdowns)

    _build_recall(wb.create_sheet("Recall Risks"), payload)

    qc_cols = payload["qc_columns"]
    _data_sheet(wb.create_sheet("Reconciliation QC"),
                "Reconciliation and acceptance controls",
                "Publication fails closed on FAIL. Value tolerance is $0.01; volume tolerance is 0.000001. "
                "WARN is retained for documented non-blocking source limitations.",
                qc_cols, payload["reconciliation_qc"], QC_WIDTHS, QC_FMT,
                cond_col=_index_of(qc_cols, "status"))

    _build_lineage(wb.create_sheet("Source Lineage"), payload)

    building = output_path.with_suffix(output_path.suffix + ".building")
    if building.exists():
        building.unlink()
    wb.save(building)
    size = building.stat().st_size
    if size >= MAX_WORKBOOK_BYTES:
        building.unlink(missing_ok=True)
        raise RuntimeError(f"Workbook is {size:,} bytes; limit is {MAX_WORKBOOK_BYTES:,} bytes.")
    os.replace(building, output_path)
    return output_path
