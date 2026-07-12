#!/usr/bin/env python3
"""Build the governed prediction-audit SQLite authority.

This is a review-only reconstruction of the current production routing state.
It does not alter mapping, guard, or terminal-routing semantics.  The database
is built at a temporary path, fully reconciled, and published with os.replace
only after every required check passes.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sqlite3
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator, Sequence

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config import settings as cfg  # noqa: E402

EXCEL_MAX_DATA_ROWS = 1_048_575
CHUNK_ROWS = 40_000
VALUE_TOLERANCE = 0.01
VOLUME_TOLERANCE = 1e-6
DIMENSION_UNMAPPED = "<Unmapped>"
DIMENSION_ALL = "<All>"

MRI_COMPATIBLE_RE = re.compile(r"\bmri[\s-]*(?:compatible|compatibility|conditional|safe)\b", re.I)
MRI_IMAGING_RE = re.compile(r"\b(?:mri|magnetic\s+resonance)\b.*\b(?:scanner|imaging|coil|magnet|tesla)\b|\b(?:scanner|imaging|coil|magnet|tesla)\b.*\b(?:mri|magnetic\s+resonance)\b", re.I)
PERIOPERATIVE_RE = re.compile(r"\b(?:surg(?:ery|ical)|operating\s+room|operating\s+table|patient\s+table|an?esthesia|perioperative|endoscop|laparoscop)\b", re.I)
DENTAL_RE = re.compile(r"\b(?:dental|dentist|orthodont|endodont|root\s+canal)\b", re.I)
ACCESSORY_RE = re.compile(r"\b(?:accessor(?:y|ies)|spare\s+part|replacement\s+part)\b", re.I)

REQUIRED_COLUMNS = {"Detailed_Product", "Total_Value_USD", "Quantity"}
FACT_COLUMNS = [
    "run_id", "output_file_id", "source_row_id", "original_unique_id", "source_row_number",
    "country", "fiscal_year", "detailed_product", "manufacturer", "family", "product",
    "segment", "sub_segment", "match_tier", "reference_status", "scope_flag", "qa_status",
    "output_tier", "primary_reason", "removal_stage_id", "raw_value_usd", "raw_volume",
    "value_usd", "volume", "value_numeric_status", "volume_numeric_status", "mri_risk",
    "mri_actual_imaging", "mri_perioperative_signal", "nonstandard_tier",
    "independent_surgical_signal", "negative_conflict_group", "generic_token_risk",
    "date_month_token_risk", "apt_march_rule_risk", "ophthalmic_imaging_conflict_risk",
    "extended_false_positive_risk", "vector_auto_mapping_status", "master_validation_status",
    "source_text_hash",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def text(value: object) -> str:
    if value is None:
        return ""
    result = str(value).strip()
    return "" if result.lower() == "nan" else result


def truthy(value: object) -> int:
    return int(text(value).lower() in {"1", "true", "yes", "y"})


def parse_number(value: object) -> tuple[float | None, str]:
    raw = text(value)
    if not raw:
        return None, "missing"
    try:
        parsed = float(raw.replace(",", ""))
    except (TypeError, ValueError):
        return None, "invalid"
    return (parsed, "valid") if math.isfinite(parsed) else (None, "invalid")


def source_text(row: dict[str, object]) -> str:
    return " | ".join(
        text(row.get(column))
        for column in ("Detailed_Product", "Importer", "Exporter", "Manufacturer", "Family", "Product_V0")
    )


def reference_is_valid(value: object) -> bool:
    return text(value).lower() in {"valid", "y", "yes", "true", "1", "reference-valid"}


def terminal_tier(row: dict[str, object]) -> str:
    explicit = text(row.get("Output_Tier"))
    aliases = {
        "Trusted_Dashboard": "Trusted",
        "Review_Queue": "Review",
        "Excluded_Unmapped": "Excluded",
    }
    if explicit in {"Trusted", "Review", "Excluded"}:
        return explicit
    if explicit in aliases:
        return aliases[explicit]
    if text(row.get("Dash_Include")).upper() == "Y":
        return "Trusted"
    qa = text(row.get("QA_Status")).lower()
    return "Review" if qa.startswith(("review", "audit")) else "Excluded"


def derive_reason(row: dict[str, object], tier: str) -> tuple[str, str]:
    qa = text(row.get("QA_Status"))
    negative = text(row.get("Negative_Conflict_Group"))
    scope = text(row.get("Scope_Flag"))
    reference = text(row.get("Reference_Key_Status") or row.get("Ref_Valid"))
    if tier == "Trusted":
        return "trusted_reference_valid_surgical", "S13_TERMINAL_ROUTING"
    if truthy(row.get("Ophthalmic_Imaging_Conflict_Risk")) or "ophthalmic" in qa.lower() or "imaging" in qa.lower():
        return "ophthalmic_imaging_conflict", "S12_REMAP_GUARDS"
    if reference and not reference_is_valid(reference):
        return "reference_tuple_invalid", "S07_REFERENCE_VALIDATION"
    if scope:
        return "scope_exclusion", "S08_SCOPE_WHITELIST"
    if negative:
        return f"negative_conflict:{negative}", "S04_CATEGORY_FALLBACK"
    if any(truthy(row.get(c)) for c in ("Generic_Token_Risk", "Date_Month_Token_Risk", "APT_March_Rule_Risk")):
        return "generic_or_token_anomaly", "S09_GENERIC_ANOMALY"
    if truthy(row.get("Extended_False_Positive_Risk")):
        return "extended_hs_false_positive_risk", "S10_EXTENDED_HS"
    if qa:
        return qa, "S13_TERMINAL_ROUTING"
    return ("review_required" if tier == "Review" else "excluded_no_accepted_candidate"), "S13_TERMINAL_ROUTING"


def sha256_file(path: Path, block_size: int = 8 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while block := handle.read(block_size):
            digest.update(block)
    return digest.hexdigest()


def count_delimited_rows(path: Path) -> int:
    line_count = 0
    last_byte = b""
    with path.open("rb") as handle:
        while block := handle.read(8 * 1024 * 1024):
            line_count += block.count(b"\n")
            last_byte = block[-1:]
    if last_byte and last_byte != b"\n":
        line_count += 1
    return max(line_count - 1, 0)


def git_commit() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "HEAD"], cwd=ROOT, text=True, stderr=subprocess.DEVNULL
        ).strip()
    except Exception:
        return "working-tree"


def validate_registry(registry: dict[str, object]) -> None:
    stage_ids = [stage["stage_id"] for stage in registry["stages"]]
    if len(stage_ids) != len(set(stage_ids)):
        raise RuntimeError("Duplicate stage_id in prediction rule registry")
    required_stages = {f"S{i:02d}_" for i in range(15)}
    if not all(any(stage_id.startswith(prefix) for stage_id in stage_ids) for prefix in required_stages):
        raise RuntimeError("Registry must cover S00 through S14")
    rule_ids = [rule["rule_id"] for rule in registry["rules"]]
    if len(rule_ids) != len(set(rule_ids)):
        raise RuntimeError("Duplicate rule_id in prediction rule registry")
    for rule in registry["rules"]:
        missing = {
            "stage_id", "rule_id", "version", "execution_order", "input_population",
            "predicate_expression", "outcome_expression", "reason_precedence", "row_continues",
        } - set(rule)
        if missing:
            raise RuntimeError(f"Registry rule {rule.get('rule_id')} missing {sorted(missing)}")
        if rule["stage_id"] not in stage_ids:
            raise RuntimeError(f"Registry rule points to unknown stage: {rule['rule_id']}")


def initialize_db(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        path.unlink()
    connection = sqlite3.connect(path)
    connection.execute("PRAGMA foreign_keys=ON")
    connection.execute("PRAGMA journal_mode=DELETE")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("PRAGMA temp_store=FILE")
    connection.execute("PRAGMA cache_size=-250000")
    connection.executescript("""
    CREATE TABLE run (
      run_id TEXT PRIMARY KEY,
      started_at TEXT NOT NULL,
      completed_at TEXT,
      registry_version TEXT NOT NULL,
      fixed_seed INTEGER NOT NULL,
      policy TEXT NOT NULL CHECK(policy='review_only'),
      status TEXT NOT NULL,
      code_commit TEXT NOT NULL,
      notes TEXT NOT NULL
    );
    CREATE TABLE rule_registry_stage (
      stage_id TEXT PRIMARY KEY,
      version TEXT NOT NULL,
      registry_version TEXT NOT NULL,
      execution_order INTEGER NOT NULL UNIQUE,
      documentation_label TEXT NOT NULL,
      stage_type TEXT NOT NULL,
      input_population TEXT NOT NULL,
      predicate_description TEXT NOT NULL,
      outcomes_json TEXT NOT NULL,
      reason_precedence_json TEXT NOT NULL,
      row_continues_json TEXT NOT NULL
    );
    CREATE TABLE rule_registry_rule (
      rule_id TEXT PRIMARY KEY,
      stage_id TEXT NOT NULL,
      version TEXT NOT NULL,
      execution_order INTEGER NOT NULL,
      documentation_label TEXT NOT NULL,
      input_population TEXT NOT NULL,
      predicate_expression TEXT NOT NULL,
      outcome_expression TEXT NOT NULL,
      reason_precedence_json TEXT NOT NULL,
      row_continues_json TEXT NOT NULL,
      FOREIGN KEY(stage_id) REFERENCES rule_registry_stage(stage_id)
    );
    CREATE TABLE source_file (
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      output_label TEXT NOT NULL,
      country TEXT NOT NULL,
      fiscal_year TEXT NOT NULL,
      source_path TEXT NOT NULL,
      complete_source_path TEXT,
      source_format TEXT NOT NULL,
      ingestion_mode TEXT NOT NULL,
      completeness_basis TEXT NOT NULL,
      expected_rows INTEGER NOT NULL,
      observed_rows INTEGER NOT NULL,
      source_sha256 TEXT NOT NULL,
      complete_source_sha256 TEXT,
      source_bytes INTEGER NOT NULL,
      complete_source_bytes INTEGER,
      is_complete INTEGER NOT NULL CHECK(is_complete IN (0,1)),
      transaction_count INTEGER,
      value_usd REAL,
      volume REAL,
      missing_value_count INTEGER,
      invalid_value_count INTEGER,
      missing_volume_count INTEGER,
      invalid_volume_count INTEGER,
      PRIMARY KEY(run_id, output_file_id),
      FOREIGN KEY(run_id) REFERENCES run(run_id)
    );
    CREATE TABLE row_fact (
      row_fact_id INTEGER PRIMARY KEY,
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      source_row_id TEXT NOT NULL,
      original_unique_id TEXT,
      source_row_number INTEGER NOT NULL,
      country TEXT NOT NULL,
      fiscal_year TEXT NOT NULL,
      detailed_product TEXT,
      manufacturer TEXT,
      family TEXT,
      product TEXT,
      segment TEXT,
      sub_segment TEXT,
      match_tier TEXT,
      reference_status TEXT,
      scope_flag TEXT,
      qa_status TEXT,
      output_tier TEXT NOT NULL CHECK(output_tier IN ('Trusted','Review','Excluded')),
      primary_reason TEXT NOT NULL,
      removal_stage_id TEXT NOT NULL,
      raw_value_usd TEXT,
      raw_volume TEXT,
      value_usd REAL,
      volume REAL,
      value_numeric_status TEXT NOT NULL CHECK(value_numeric_status IN ('valid','missing','invalid')),
      volume_numeric_status TEXT NOT NULL CHECK(volume_numeric_status IN ('valid','missing','invalid')),
      mri_risk INTEGER NOT NULL CHECK(mri_risk IN (0,1)),
      mri_actual_imaging INTEGER NOT NULL CHECK(mri_actual_imaging IN (0,1)),
      mri_perioperative_signal INTEGER NOT NULL CHECK(mri_perioperative_signal IN (0,1)),
      nonstandard_tier INTEGER NOT NULL CHECK(nonstandard_tier IN (0,1)),
      independent_surgical_signal INTEGER NOT NULL CHECK(independent_surgical_signal IN (0,1)),
      negative_conflict_group TEXT,
      generic_token_risk INTEGER NOT NULL CHECK(generic_token_risk IN (0,1)),
      date_month_token_risk INTEGER NOT NULL CHECK(date_month_token_risk IN (0,1)),
      apt_march_rule_risk INTEGER NOT NULL CHECK(apt_march_rule_risk IN (0,1)),
      ophthalmic_imaging_conflict_risk INTEGER NOT NULL CHECK(ophthalmic_imaging_conflict_risk IN (0,1)),
      extended_false_positive_risk INTEGER NOT NULL CHECK(extended_false_positive_risk IN (0,1)),
      vector_auto_mapping_status TEXT,
      master_validation_status TEXT,
      source_text_hash TEXT NOT NULL,
      UNIQUE(run_id, output_file_id, source_row_id),
      FOREIGN KEY(run_id, output_file_id) REFERENCES source_file(run_id, output_file_id),
      FOREIGN KEY(removal_stage_id) REFERENCES rule_registry_stage(stage_id)
    );
    CREATE TABLE row_stage_state (
      row_fact_id INTEGER NOT NULL,
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      source_row_id TEXT NOT NULL,
      stage_id TEXT NOT NULL,
      stage_order INTEGER NOT NULL,
      rule_id TEXT NOT NULL,
      rule_version TEXT NOT NULL,
      candidate_status TEXT NOT NULL CHECK(candidate_status IN ('Eligible','Hit','Suppressed','Released','Missed','Recovered')),
      outcome TEXT NOT NULL,
      reason TEXT NOT NULL,
      row_continues INTEGER NOT NULL CHECK(row_continues IN (0,1)),
      PRIMARY KEY(row_fact_id, stage_id),
      FOREIGN KEY(row_fact_id) REFERENCES row_fact(row_fact_id),
      FOREIGN KEY(run_id, output_file_id) REFERENCES source_file(run_id, output_file_id),
      FOREIGN KEY(stage_id) REFERENCES rule_registry_stage(stage_id),
      FOREIGN KEY(rule_id) REFERENCES rule_registry_rule(rule_id)
    ) WITHOUT ROWID;
    CREATE TABLE rule_hit (
      rule_hit_id INTEGER PRIMARY KEY,
      row_fact_id INTEGER NOT NULL,
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      source_row_id TEXT NOT NULL,
      stage_id TEXT NOT NULL,
      rule_id TEXT NOT NULL,
      rule_version TEXT NOT NULL,
      reason TEXT NOT NULL,
      hit_kind TEXT NOT NULL CHECK(hit_kind IN ('primary','secondary')),
      is_additive INTEGER NOT NULL CHECK(is_additive IN (0,1)),
      FOREIGN KEY(row_fact_id) REFERENCES row_fact(row_fact_id),
      FOREIGN KEY(run_id, output_file_id) REFERENCES source_file(run_id, output_file_id),
      FOREIGN KEY(stage_id) REFERENCES rule_registry_stage(stage_id),
      FOREIGN KEY(rule_id) REFERENCES rule_registry_rule(rule_id)
    );
    CREATE TABLE review_label (
      review_label_id INTEGER PRIMARY KEY,
      row_fact_id INTEGER NOT NULL,
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      source_row_id TEXT NOT NULL,
      sample_type TEXT NOT NULL,
      sample_stratum TEXT NOT NULL,
      target_category TEXT NOT NULL,
      inclusion_probability REAL,
      sample_weight REAL,
      fixed_seed INTEGER NOT NULL,
      sample_rank INTEGER NOT NULL,
      evidence TEXT NOT NULL,
      shadow_recommendation TEXT NOT NULL,
      production_changed INTEGER NOT NULL DEFAULT 0 CHECK(production_changed=0),
      surgical_relevance TEXT,
      mapping_correctness TEXT,
      corrected_manufacturer TEXT,
      corrected_family TEXT,
      corrected_product TEXT,
      corrected_segment TEXT,
      corrected_sub_segment TEXT,
      reviewer_rationale TEXT,
      reviewer TEXT,
      reviewed_at TEXT,
      adjudicator TEXT,
      adjudicated_at TEXT,
      disposition TEXT NOT NULL DEFAULT 'Pending',
      UNIQUE(run_id, output_file_id, source_row_id),
      FOREIGN KEY(row_fact_id) REFERENCES row_fact(row_fact_id)
    );
    CREATE TABLE recall_risk_inventory (
      inventory_id INTEGER PRIMARY KEY,
      row_fact_id INTEGER NOT NULL,
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      source_row_id TEXT NOT NULL,
      risk_type TEXT NOT NULL,
      current_output_tier TEXT NOT NULL,
      evidence TEXT NOT NULL,
      recommendation TEXT NOT NULL,
      UNIQUE(run_id, output_file_id, source_row_id, risk_type),
      FOREIGN KEY(row_fact_id) REFERENCES row_fact(row_fact_id)
    );
    CREATE TABLE funnel_cube (
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      output_label TEXT NOT NULL,
      funnel_type TEXT NOT NULL,
      stage_id TEXT NOT NULL,
      stage_order INTEGER NOT NULL,
      stage_label TEXT NOT NULL,
      rule_version TEXT NOT NULL,
      candidate_status TEXT NOT NULL,
      outcome TEXT NOT NULL,
      transaction_count INTEGER NOT NULL,
      value_usd REAL,
      volume REAL,
      missing_value_count INTEGER NOT NULL,
      invalid_value_count INTEGER NOT NULL,
      missing_volume_count INTEGER NOT NULL,
      invalid_volume_count INTEGER NOT NULL,
      weighted_asp REAL,
      previous_stage_transaction_count INTEGER NOT NULL,
      previous_stage_value_usd REAL,
      previous_stage_volume REAL,
      filtered_transaction_count INTEGER NOT NULL,
      filtered_value_usd REAL,
      filtered_volume REAL,
      filtered_value_pct REAL,
      PRIMARY KEY(run_id, output_file_id, stage_id, candidate_status, outcome)
    );
    CREATE TABLE removal_cube (
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      output_label TEXT NOT NULL,
      reason_kind TEXT NOT NULL CHECK(reason_kind IN ('Primary','Secondary')),
      is_additive INTEGER NOT NULL CHECK(is_additive IN (0,1)),
      stage_id TEXT NOT NULL,
      rule_id TEXT NOT NULL,
      rule_version TEXT NOT NULL,
      outcome TEXT NOT NULL,
      reason TEXT NOT NULL,
      grouping_level TEXT NOT NULL,
      grouping_id TEXT NOT NULL,
      manufacturer TEXT NOT NULL,
      family TEXT NOT NULL,
      product TEXT NOT NULL,
      transaction_count INTEGER NOT NULL,
      value_usd REAL,
      volume REAL,
      missing_value_count INTEGER NOT NULL,
      invalid_value_count INTEGER NOT NULL,
      missing_volume_count INTEGER NOT NULL,
      invalid_volume_count INTEGER NOT NULL,
      weighted_asp REAL,
      previous_stage_transaction_count INTEGER,
      previous_stage_value_usd REAL,
      previous_stage_volume REAL,
      filtered_transaction_count INTEGER,
      filtered_value_usd REAL,
      filtered_volume REAL,
      filtered_value_pct REAL
    );
    CREATE TABLE reconciliation_qc (
      qc_id INTEGER PRIMARY KEY,
      run_id TEXT NOT NULL,
      output_file_id TEXT NOT NULL,
      check_name TEXT NOT NULL,
      observed TEXT NOT NULL,
      expected TEXT NOT NULL,
      value_delta REAL,
      volume_delta REAL,
      status TEXT NOT NULL CHECK(status IN ('PASS','FAIL','WARN')),
      evidence TEXT NOT NULL,
      checked_at TEXT NOT NULL
    );
    CREATE TABLE baseline_manifest (
      manifest_id INTEGER PRIMARY KEY,
      run_id TEXT NOT NULL,
      artifact_type TEXT NOT NULL,
      path TEXT NOT NULL,
      sha256 TEXT NOT NULL,
      bytes INTEGER NOT NULL,
      transaction_count INTEGER,
      value_usd REAL,
      volume REAL,
      UNIQUE(run_id, path)
    );
    CREATE TABLE artifact_manifest (
      artifact_id INTEGER PRIMARY KEY,
      run_id TEXT NOT NULL,
      artifact_type TEXT NOT NULL,
      path TEXT NOT NULL,
      sha256 TEXT NOT NULL,
      bytes INTEGER NOT NULL,
      generated_at TEXT NOT NULL,
      UNIQUE(run_id, path)
    );
    CREATE TABLE independent_qc_finding (
      finding_id INTEGER PRIMARY KEY,
      run_id TEXT NOT NULL,
      severity TEXT NOT NULL,
      finding TEXT NOT NULL,
      evidence TEXT NOT NULL,
      stage_id TEXT,
      rule_id TEXT,
      file_path TEXT,
      required_action TEXT NOT NULL,
      owner_response TEXT,
      retest_result TEXT,
      status TEXT NOT NULL DEFAULT 'Open'
    );
    """)
    return connection


def load_registry(connection: sqlite3.Connection, registry: dict[str, object]) -> dict[str, dict[str, object]]:
    stage_rows = []
    for stage in registry["stages"]:
        stage_rows.append((
            stage["stage_id"], stage["version"], registry["registry_version"], stage["execution_order"],
            stage["documentation_label"], stage["stage_type"], stage["input_population"],
            stage["predicate_description"], json.dumps(stage["outcomes"], separators=(",", ":")),
            json.dumps(stage["reason_precedence"], separators=(",", ":")),
            json.dumps(stage["row_continues"], separators=(",", ":")),
        ))
    connection.executemany("INSERT INTO rule_registry_stage VALUES (?,?,?,?,?,?,?,?,?,?,?)", stage_rows)
    rule_rows = []
    for rule in registry["rules"]:
        rule_rows.append((
            rule["rule_id"], rule["stage_id"], rule["version"], rule["execution_order"],
            rule["documentation_label"], rule["input_population"], rule["predicate_expression"],
            rule["outcome_expression"], json.dumps(rule["reason_precedence"], separators=(",", ":")),
            json.dumps(rule["row_continues"], separators=(",", ":")),
        ))
    connection.executemany("INSERT INTO rule_registry_rule VALUES (?,?,?,?,?,?,?,?,?,?)", rule_rows)
    return {str(rule["rule_id"]): rule for rule in registry["rules"]}


def iter_legacy_xlsx(path: Path) -> tuple[int, Iterator[list[dict[str, object]]], list[str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "RawData" not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError(f"FAIL CLOSED: RawData sheet missing: {path}")
    worksheet = workbook["RawData"]
    observed = worksheet.max_row - 1
    if observed >= EXCEL_MAX_DATA_ROWS:
        workbook.close()
        raise RuntimeError(f"FAIL CLOSED: Excel source is capped/truncated ({observed:,} data rows): {path}")
    headers = [text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        workbook.close()
        raise RuntimeError(f"FAIL CLOSED: required columns missing from {path}: {sorted(missing)}")

    def chunks() -> Iterator[list[dict[str, object]]]:
        batch: list[dict[str, object]] = []
        try:
            for values in worksheet.iter_rows(min_row=2, values_only=True):
                batch.append(dict(zip(headers, values)))
                if len(batch) >= CHUNK_ROWS:
                    yield batch
                    batch = []
            if batch:
                yield batch
        finally:
            workbook.close()

    return observed, chunks(), headers


def iter_current_csv(path: Path) -> tuple[int, Iterator[list[dict[str, object]]], list[str]]:
    observed = count_delimited_rows(path)
    headers = list(pd.read_csv(path, dtype=str, nrows=0).columns)
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise RuntimeError(f"FAIL CLOSED: required columns missing from {path}: {sorted(missing)}")

    def chunks() -> Iterator[list[dict[str, object]]]:
        for frame in pd.read_csv(
            path, dtype=str, keep_default_na=False, chunksize=CHUNK_ROWS, low_memory=False
        ):
            yield frame.to_dict("records")

    return observed, chunks(), headers


def stage_state(
    fact_id: int,
    run_id: str,
    output_id: str,
    source_row_id: str,
    stage_id: str,
    stage_order: int,
    rule_id: str,
    rule_version: str,
    candidate_status: str,
    outcome: str,
    reason: str,
    continues: int,
) -> tuple[object, ...]:
    return (
        fact_id, run_id, output_id, source_row_id, stage_id, stage_order, rule_id,
        rule_version, candidate_status, outcome, reason, continues,
    )


def build_stage_states(
    fact_id: int,
    run_id: str,
    output_id: str,
    source_row_id: str,
    row: dict[str, object],
    tier: str,
    primary_reason: str,
    rules: dict[str, dict[str, object]],
) -> list[tuple[object, ...]]:
    description = source_text(row)
    match_tier = text(row.get("Match_Tier")).lower()
    mapped = any(text(row.get(c)) for c in ("Manufacturer", "Family", "Segment", "Sub-segment", "Product_V0"))
    negative = text(row.get("Negative_Conflict_Group"))
    scope = text(row.get("Scope_Flag"))
    reference = text(row.get("Reference_Key_Status") or row.get("Ref_Valid"))

    states: list[tuple[object, ...]] = []

    def add(rule_id: str, candidate: str, outcome: str, reason: str, continues: int = 1) -> None:
        rule = rules[rule_id]
        states.append(stage_state(
            fact_id, run_id, output_id, source_row_id, str(rule["stage_id"]),
            int(rule["execution_order"]), rule_id, str(rule["version"]), candidate,
            outcome, reason, continues,
        ))

    add("S00_COMPLETE_SOURCE", "Eligible", "Passed", "complete_source_loaded")
    add("S01_MATCH_ALL_HS4", "Eligible", "Eligible", "hs_eligible")
    if DENTAL_RE.search(description):
        recovered = truthy(row.get("Independent_Surgical_Signal")) == 1
        add(
            "S02_DENTAL_SIGNAL", "Recovered" if recovered else "Suppressed",
            "Recovered" if recovered else "Suppressed",
            "dental_with_independent_surgical_signal" if recovered else "dental_only",
        )
    if "family" in match_tier:
        add("S03_FAMILY_CANDIDATE", "Hit", "Hit", "family_match")
    if "category" in match_tier or negative or ACCESSORY_RE.search(description):
        if negative:
            add("S04_CATEGORY_OR_CONFLICT", "Suppressed", "Suppressed", f"negative_conflict:{negative}")
        elif ACCESSORY_RE.search(description) and "category" not in match_tier:
            add("S04_CATEGORY_OR_CONFLICT", "Missed", "Review", "accessory_only")
        else:
            add("S04_CATEGORY_OR_CONFLICT", "Hit", "Hit", "category_match")
    if "manufacturer" in match_tier:
        add(
            "S05_MANUFACTURER_CANDIDATE", "Hit" if tier == "Trusted" else "Missed",
            "Hit" if tier == "Trusted" else "Review", "manufacturer_fallback",
        )
    if mapped:
        add("S06_CANONICAL_DIMENSIONS", "Hit", "Hit", "standardized")
        valid = reference_is_valid(reference)
        add(
            "S07_REFERENCE_TUPLE", "Hit" if valid else "Missed", "Hit" if valid else "Review",
            "reference_tuple_valid" if valid else ("reference_tuple_invalid" if reference else "reference_tuple_missing"),
        )
    independent_signal = truthy(row.get("Independent_Surgical_Signal")) == 1
    if scope or independent_signal:
        add(
            "S08_SCOPE_CONTROL", "Recovered" if independent_signal else "Suppressed",
            "Recovered" if independent_signal else "Suppressed",
            "independent_surgical_recovery" if independent_signal else "scope_exclusion",
        )
    if any(
        truthy(row.get(c))
        for c in ("Generic_Token_Risk", "Date_Month_Token_Risk", "APT_March_Rule_Risk")
    ) or text(row.get("High_Risk_Token")):
        add("S09_TOKEN_ANOMALY", "Suppressed", "Review" if tier == "Review" else "Suppressed", "generic_or_token_anomaly")
    if "extended" in match_tier or truthy(row.get("Extended_False_Positive_Risk")):
        risky = truthy(row.get("Extended_False_Positive_Risk")) == 1
        add(
            "S10_EXTENDED_HS_CONTROL", "Suppressed" if risky else "Hit",
            "Suppressed" if risky else "Hit",
            "extended_hs_false_positive_risk" if risky else "extended_hs_supported",
        )
    if match_tier == "hs_prior":
        add(
            "S11_HS_PRIOR_RECOVERY", "Recovered" if tier == "Trusted" else "Missed",
            "Recovered" if tier == "Trusted" else "Review", "hs_prior_recovery",
        )
    add(
        "S12_FINAL_GUARD",
        "Hit" if tier == "Trusted" else ("Missed" if tier == "Review" else "Suppressed"),
        "Hit" if tier == "Trusted" else ("Review" if tier == "Review" else "Suppressed"),
        primary_reason,
    )
    add(
        "S13_EXACTLY_ONE_ROUTE",
        "Eligible" if tier == "Trusted" else ("Missed" if tier == "Review" else "Suppressed"),
        tier, primary_reason, 0,
    )
    return states


def build_rule_hits(
    fact_id: int,
    run_id: str,
    output_id: str,
    source_row_id: str,
    row: dict[str, object],
    tier: str,
    primary_reason: str,
    removal_stage: str,
    mri_risk: int,
    nonstandard: int,
    rules: dict[str, dict[str, object]],
) -> list[tuple[object, ...]]:
    hits: list[tuple[object, ...]] = []

    def add(rule_id: str, reason: str, kind: str = "secondary", additive: int = 0, stage_id: str | None = None) -> None:
        rule = rules[rule_id]
        hits.append((
            fact_id, run_id, output_id, source_row_id, stage_id or rule["stage_id"], rule_id,
            rule["version"], reason, kind, additive,
        ))

    if tier != "Trusted":
        stage_rule = {
            "S04_CATEGORY_FALLBACK": "S04_CATEGORY_OR_CONFLICT",
            "S07_REFERENCE_VALIDATION": "S07_REFERENCE_TUPLE",
            "S08_SCOPE_WHITELIST": "S08_SCOPE_CONTROL",
            "S09_GENERIC_ANOMALY": "S09_TOKEN_ANOMALY",
            "S10_EXTENDED_HS": "S10_EXTENDED_HS_CONTROL",
            "S12_REMAP_GUARDS": "S12_FINAL_GUARD",
            "S13_TERMINAL_ROUTING": "S13_EXACTLY_ONE_ROUTE",
        }
        add(stage_rule.get(removal_stage, "S13_EXACTLY_ONE_ROUTE"), primary_reason, "primary", 1, removal_stage)
    description = source_text(row)
    if mri_risk:
        add("S12_MRI_COMPATIBLE_RISK", "mri_compatible_recall_risk")
    if nonstandard:
        rule_id = "S11_HS_PRIOR_RECOVERY" if text(row.get("Match_Tier")).lower() == "hs_prior" else "S05_MANUFACTURER_CANDIDATE"
        add(rule_id, f"nonstandard_tier_retained:{text(row.get('Match_Tier'))}")
    if text(row.get("Negative_Conflict_Group")):
        add("S04_CATEGORY_OR_CONFLICT", f"negative_conflict:{text(row.get('Negative_Conflict_Group'))}")
    if text(row.get("Scope_Flag")):
        add("S08_SCOPE_CONTROL", f"scope_flag:{text(row.get('Scope_Flag'))}")
    if any(truthy(row.get(c)) for c in ("Generic_Token_Risk", "Date_Month_Token_Risk", "APT_March_Rule_Risk")):
        add("S09_TOKEN_ANOMALY", "generic_or_token_anomaly")
    if truthy(row.get("Ophthalmic_Imaging_Conflict_Risk")):
        add("S12_FINAL_GUARD", "ophthalmic_imaging_conflict")
    if truthy(row.get("Extended_False_Positive_Risk")):
        add("S10_EXTENDED_HS_CONTROL", "extended_hs_false_positive_risk")
    if DENTAL_RE.search(description):
        add("S02_DENTAL_SIGNAL", "dental_cue")
    return hits


def build_inventory_rows(
    fact_id: int,
    run_id: str,
    output_id: str,
    source_row_id: str,
    tier: str,
    reference_status: str,
    match_tier: str,
    description: str,
    mri_risk: int,
    mri_actual: int,
    mri_perioperative: int,
    nonstandard: int,
) -> list[tuple[object, ...]]:
    rows: list[tuple[object, ...]] = []
    if mri_risk:
        rows.append((
            fact_id, run_id, output_id, source_row_id, "MRI compatible recall risk", tier,
            "MRI-compatible regex matched the governed source-text composite",
            "Human review only; no automatic promotion in this release",
        ))
    if mri_actual:
        rows.append((
            fact_id, run_id, output_id, source_row_id, "MRI actual imaging-system conflict", tier,
            "MRI and scanner/imaging/coil/magnet/tesla cues co-occur",
            "Keep separate from perioperative MRI-compatible equipment and confirm Excluded routing",
        ))
    if mri_perioperative:
        rows.append((
            fact_id, run_id, output_id, source_row_id, "MRI perioperative surgical signal", tier,
            "MRI-compatible wording co-occurs with surgical/perioperative equipment cues",
            "Consider Review only after adjudication and governed reference evidence",
        ))
    if mri_risk and tier == "Trusted" and reference_is_valid(reference_status):
        rows.append((
            fact_id, run_id, output_id, source_row_id, "MRI named reference-valid family retained Trusted", tier,
            "MRI-compatible wording is already reference-valid and Trusted under current production semantics",
            "Retain Trusted unless human review identifies a true imaging-system conflict",
        ))
    if nonstandard:
        rows.append((
            fact_id, run_id, output_id, source_row_id, f"Non-standard tier: {match_tier}", tier,
            f"Match_Tier={match_tier}; mapped dimensions retained in row_fact",
            "Retain in SQLite and review evidence before any future Trusted promotion",
        ))
    return rows


def metrics_sql(alias: str = "") -> str:
    prefix = f"{alias}." if alias else ""
    return f"""
      COUNT(*) AS transaction_count,
      SUM({prefix}value_usd) AS value_usd,
      SUM({prefix}volume) AS volume,
      SUM(CASE WHEN {prefix}value_numeric_status='missing' THEN 1 ELSE 0 END) AS missing_value_count,
      SUM(CASE WHEN {prefix}value_numeric_status='invalid' THEN 1 ELSE 0 END) AS invalid_value_count,
      SUM(CASE WHEN {prefix}volume_numeric_status='missing' THEN 1 ELSE 0 END) AS missing_volume_count,
      SUM(CASE WHEN {prefix}volume_numeric_status='invalid' THEN 1 ELSE 0 END) AS invalid_volume_count,
      CASE WHEN SUM({prefix}volume) != 0 THEN SUM({prefix}value_usd)/SUM({prefix}volume) END AS weighted_asp
    """


def ingest_output(
    connection: sqlite3.Connection,
    run_id: str,
    spec: dict[str, object],
    rules: dict[str, dict[str, object]],
) -> None:
    path = ROOT / str(spec["path"])
    if not path.exists():
        raise RuntimeError(f"FAIL CLOSED: source missing: {path}")
    expected = int(spec["expected_rows"])
    mode = str(spec["ingestion_mode"])
    complete_path = ROOT / str(spec["complete_source_path"]) if spec.get("complete_source_path") else None
    if mode == "governed_uncapped_legacy_excel_migration":
        observed, chunks, _headers = iter_legacy_xlsx(path)
        complete_path = path
        source_format = "xlsx"
        completeness_basis = "Sub-cap RawData worksheet; governed one-time legacy migration"
    elif mode == "complete_csv_current_remap":
        if complete_path is None or not complete_path.exists():
            raise RuntimeError(f"FAIL CLOSED: complete source missing: {complete_path}")
        complete_rows = count_delimited_rows(complete_path)
        if complete_rows != expected:
            raise RuntimeError(f"FAIL CLOSED: complete source count {complete_rows:,} != expected {expected:,}")
        observed, chunks, _headers = iter_current_csv(path)
        if observed != complete_rows:
            raise RuntimeError(f"FAIL CLOSED: mapped CSV count {observed:,} != complete source {complete_rows:,}")
        source_format = "csv"
        completeness_basis = "Complete mapped CSV reconciled to independently counted immutable raw CSV"
    else:
        raise RuntimeError(f"Unsupported ingestion_mode: {mode}")
    if observed != expected:
        raise RuntimeError(f"FAIL CLOSED: {spec['output_label']} observed {observed:,} != expected {expected:,}")

    source_hash = sha256_file(path)
    complete_hash = sha256_file(complete_path) if complete_path else None
    output_id = str(spec["output_file_id"])
    connection.execute(
        """INSERT INTO source_file (
          run_id,output_file_id,output_label,country,fiscal_year,source_path,complete_source_path,
          source_format,ingestion_mode,completeness_basis,expected_rows,observed_rows,source_sha256,
          complete_source_sha256,source_bytes,complete_source_bytes,is_complete
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",
        (
            run_id, output_id, str(spec["output_label"]), str(spec["country"]), str(spec["fiscal_year"]),
            str(path.resolve()), str(complete_path.resolve()) if complete_path else None,
            source_format, mode, completeness_basis, expected, observed, source_hash, complete_hash,
            path.stat().st_size, complete_path.stat().st_size if complete_path else None,
        ),
    )
    connection.commit()

    fact_sql = f"INSERT INTO row_fact ({','.join(FACT_COLUMNS)}) VALUES ({','.join('?' for _ in FACT_COLUMNS)})"
    fact_id = int(connection.execute("SELECT COALESCE(MAX(row_fact_id),0) FROM row_fact").fetchone()[0])
    processed = 0
    for batch in chunks:
        facts: list[tuple[object, ...]] = []
        states: list[tuple[object, ...]] = []
        hits: list[tuple[object, ...]] = []
        inventory: list[tuple[object, ...]] = []
        for row in batch:
            processed += 1
            fact_id += 1
            source_row_id = f"{output_id}:{processed:09d}"
            original_uid = text(row.get("UniqueID"))
            description = source_text(row)
            tier = terminal_tier(row)
            reason, removal_stage = derive_reason(row, tier)
            raw_value = text(row.get("Total_Value_USD"))
            raw_volume = text(row.get("Quantity"))
            value, value_status = parse_number(raw_value)
            volume, volume_status = parse_number(raw_volume)
            match_tier = text(row.get("Match_Tier"))
            reference_status = text(row.get("Reference_Key_Status") or row.get("Ref_Valid"))
            mri_risk = int(bool(MRI_COMPATIBLE_RE.search(description)))
            mri_actual = int(bool(MRI_IMAGING_RE.search(description)))
            mri_perioperative = int(bool(mri_risk and PERIOPERATIVE_RE.search(description)))
            nonstandard = int(match_tier.lower() in {"manufacturer", "hs_prior"})
            facts.append((
                run_id, output_id, source_row_id, original_uid, processed, str(spec["country"]),
                str(spec["fiscal_year"]), text(row.get("Detailed_Product")), text(row.get("Manufacturer")),
                text(row.get("Family")), text(row.get("Product_V0")), text(row.get("Segment")),
                text(row.get("Sub-segment")), match_tier, reference_status, text(row.get("Scope_Flag")),
                text(row.get("QA_Status")), tier, reason, removal_stage, raw_value, raw_volume, value,
                volume, value_status, volume_status, mri_risk, mri_actual, mri_perioperative,
                nonstandard, truthy(row.get("Independent_Surgical_Signal")),
                text(row.get("Negative_Conflict_Group")), truthy(row.get("Generic_Token_Risk")),
                truthy(row.get("Date_Month_Token_Risk")), truthy(row.get("APT_March_Rule_Risk")),
                truthy(row.get("Ophthalmic_Imaging_Conflict_Risk")),
                truthy(row.get("Extended_False_Positive_Risk")), text(row.get("Vector_Auto_Mapping_Status")),
                text(row.get("Master_Validation_Status")),
                hashlib.sha256(description.encode("utf-8", errors="replace")).hexdigest(),
            ))
            states.extend(build_stage_states(
                fact_id, run_id, output_id, source_row_id, row, tier, reason, rules
            ))
            hits.extend(build_rule_hits(
                fact_id, run_id, output_id, source_row_id, row, tier, reason, removal_stage,
                mri_risk, nonstandard, rules,
            ))
            inventory.extend(build_inventory_rows(
                fact_id, run_id, output_id, source_row_id, tier, reference_status, match_tier,
                description, mri_risk, mri_actual, mri_perioperative, nonstandard,
            ))
        connection.executemany(fact_sql, facts)
        connection.executemany(
            "INSERT INTO row_stage_state VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", states
        )
        connection.executemany(
            """INSERT INTO rule_hit (
              row_fact_id,run_id,output_file_id,source_row_id,stage_id,rule_id,rule_version,
              reason,hit_kind,is_additive
            ) VALUES (?,?,?,?,?,?,?,?,?,?)""",
            hits,
        )
        connection.executemany(
            """INSERT OR IGNORE INTO recall_risk_inventory (
              row_fact_id,run_id,output_file_id,source_row_id,risk_type,current_output_tier,
              evidence,recommendation
            ) VALUES (?,?,?,?,?,?,?,?)""",
            inventory,
        )
        connection.commit()
        print(f"  [{spec['output_label']}] {processed:,}/{observed:,}", flush=True)
    if processed != observed:
        raise RuntimeError(f"FAIL CLOSED: processed {processed:,} != observed {observed:,} for {spec['output_label']}")
    metrics = connection.execute(
        f"SELECT {metrics_sql()} FROM row_fact WHERE run_id=? AND output_file_id=?",
        (run_id, output_id),
    ).fetchone()
    connection.execute(
        """UPDATE source_file SET transaction_count=?,value_usd=?,volume=?,missing_value_count=?,
           invalid_value_count=?,missing_volume_count=?,invalid_volume_count=?
           WHERE run_id=? AND output_file_id=?""",
        (*metrics[:7], run_id, output_id),
    )
    connection.commit()


def create_indexes_and_views(connection: sqlite3.Connection) -> None:
    connection.executescript("""
    CREATE INDEX idx_fact_output_tier ON row_fact(run_id, output_file_id, output_tier);
    CREATE INDEX idx_fact_reason ON row_fact(run_id, output_file_id, primary_reason);
    CREATE INDEX idx_fact_dimensions ON row_fact(run_id, output_file_id, manufacturer, family, product);
    CREATE INDEX idx_fact_risks ON row_fact(run_id, output_file_id, mri_risk, nonstandard_tier);
    CREATE INDEX idx_stage_lookup ON row_stage_state(run_id, output_file_id, stage_id, candidate_status, outcome);
    CREATE INDEX idx_rule_hit_lookup ON rule_hit(run_id, output_file_id, hit_kind, stage_id, reason);
    CREATE INDEX idx_inventory_lookup ON recall_risk_inventory(run_id, output_file_id, risk_type);
    CREATE INDEX idx_review_lookup ON review_label(run_id, output_file_id, sample_type, sample_rank);
    CREATE VIEW v_recall_risk_summary AS
      SELECT i.run_id, i.output_file_id, s.output_label, i.risk_type,
             i.current_output_tier, COUNT(*) AS transaction_count,
             SUM(f.value_usd) AS value_usd, SUM(f.volume) AS volume
      FROM recall_risk_inventory i
      JOIN row_fact f ON f.row_fact_id=i.row_fact_id
      JOIN source_file s ON s.run_id=i.run_id AND s.output_file_id=i.output_file_id
      GROUP BY i.run_id, i.output_file_id, s.output_label, i.risk_type, i.current_output_tier;
    CREATE VIEW v_review_samples AS
      SELECT l.*, s.output_label, f.country, f.fiscal_year, f.detailed_product,
             f.manufacturer, f.family, f.product, f.segment, f.sub_segment,
             f.match_tier, f.reference_status, f.scope_flag, f.qa_status,
             f.output_tier, f.primary_reason, f.raw_value_usd, f.raw_volume,
             f.value_usd, f.volume, f.value_numeric_status, f.volume_numeric_status
      FROM review_label l
      JOIN row_fact f ON f.row_fact_id=l.row_fact_id
      JOIN source_file s ON s.run_id=l.run_id AND s.output_file_id=l.output_file_id;
    """)
    connection.commit()


def stable_rand(seed: int, value: str) -> int:
    digest = hashlib.sha256(f"{seed}|{value}".encode("utf-8")).hexdigest()
    return int(digest[:15], 16)


def shadow_recommendation(row: sqlite3.Row) -> str:
    if row["mri_actual_imaging"]:
        return "Confirm imaging-system conflict; retain Excluded unless evidence disproves the conflict."
    if row["mri_perioperative_signal"] and row["output_tier"] != "Trusted":
        return "Review perioperative MRI-compatible evidence; promotion remains shadow-only."
    if row["nonstandard_tier"]:
        return "Validate the nonstandard manufacturer/HS-prior route against the governed reference tuple."
    if row["output_tier"] == "Trusted":
        return "Confirm current Trusted mapping and reference evidence."
    if row["output_tier"] == "Review":
        return "Adjudicate surgical relevance and mapping correctness; do not change production output."
    return "Confirm exclusion reason and test for an independently supported surgical mapping."


def build_review_samples(
    connection: sqlite3.Connection, run_id: str, fixed_seed: int
) -> None:
    connection.row_factory = sqlite3.Row
    connection.create_function("stable_rand", 2, stable_rand, deterministic=True)
    connection.execute("DELETE FROM review_label WHERE run_id=?", (run_id,))
    categories: list[tuple[str, str, str]] = [
        ("MRI signal", "mri_risk=1", "mri_risk DESC, value_usd DESC"),
        ("Nonstandard tier", "nonstandard_tier=1", "value_usd DESC"),
        ("Top-value exclusion", "output_tier='Excluded'", "value_usd DESC"),
        ("Top-volume exclusion", "output_tier='Excluded'", "volume DESC"),
        ("Reference gap", "output_tier<>'Trusted' AND (reference_status IS NULL OR trim(reference_status)='' OR lower(reference_status) NOT IN ('1','true','valid','yes','y'))", "value_usd DESC"),
        ("Scope conflict", "scope_flag IS NOT NULL AND trim(scope_flag)<>''", "value_usd DESC"),
        ("Weak category", "lower(match_tier) LIKE '%category%' AND output_tier<>'Trusted'", "value_usd DESC"),
        ("Known regression", "(mri_perioperative_signal=1 OR ophthalmic_imaging_conflict_risk=1)", "value_usd DESC"),
        ("Surgical signal not Trusted", "independent_surgical_signal=1 AND output_tier<>'Trusted'", "value_usd DESC"),
        ("Manufacturer fallback", "lower(match_tier)='manufacturer'", "value_usd DESC"),
        ("HS-prior recovery", "lower(match_tier)='hs_prior'", "value_usd DESC"),
        ("Numeric issue", "value_numeric_status<>'valid' OR volume_numeric_status<>'valid'", "value_usd DESC"),
    ]
    output_ids = [r[0] for r in connection.execute(
        "SELECT output_file_id FROM source_file WHERE run_id=? ORDER BY output_file_id", (run_id,)
    )]
    insert_sql = """INSERT INTO review_label (
      row_fact_id,run_id,output_file_id,source_row_id,sample_type,sample_stratum,
      target_category,inclusion_probability,sample_weight,fixed_seed,sample_rank,evidence,
      shadow_recommendation,production_changed
    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,0)"""
    for output_id in output_ids:
        chosen: set[int] = set()
        sample_rank = 0
        for category, predicate, ordering in categories:
            excluded = ""
            params: list[object] = [run_id, output_id]
            if chosen:
                placeholders = ",".join("?" for _ in chosen)
                excluded = f" AND row_fact_id NOT IN ({placeholders})"
                params.extend(sorted(chosen))
            row = connection.execute(
                f"SELECT * FROM row_fact WHERE run_id=? AND output_file_id=? AND ({predicate}){excluded} "
                f"ORDER BY {ordering}, stable_rand(?, source_row_id) LIMIT 1",
                (*params, fixed_seed),
            ).fetchone()
            if row is None:
                row = connection.execute(
                    f"SELECT * FROM row_fact WHERE run_id=? AND output_file_id=?{excluded} "
                    "ORDER BY CASE output_tier WHEN 'Review' THEN 0 WHEN 'Excluded' THEN 1 ELSE 2 END, "
                    "value_usd DESC, stable_rand(?, source_row_id) LIMIT 1",
                    (*params, fixed_seed),
                ).fetchone()
            if row is None:
                raise RuntimeError(f"Unable to select targeted sample {category} for {output_id}")
            chosen.add(int(row["row_fact_id"]))
            sample_rank += 1
            evidence = (
                f"Purposeful target={category}; current route={row['output_tier']}; "
                f"primary reason={row['primary_reason']}. Purposeful rows are not probability estimates."
            )
            connection.execute(insert_sql, (
                row["row_fact_id"], run_id, output_id, row["source_row_id"], "Targeted",
                category, category, None, None, fixed_seed, sample_rank, evidence,
                shadow_recommendation(row),
            ))

        targeted_exclusions = tuple(sorted(chosen))
        placeholders = ",".join("?" for _ in targeted_exclusions)
        strata = connection.execute(
            f"""SELECT output_tier,primary_reason,COUNT(*) AS n
                FROM row_fact WHERE run_id=? AND output_file_id=?
                  AND row_fact_id NOT IN ({placeholders})
                GROUP BY output_tier,primary_reason ORDER BY n DESC,output_tier,primary_reason""",
            (run_id, output_id, *targeted_exclusions),
        ).fetchall()
        if not strata:
            raise RuntimeError(f"No random sampling population remains for {output_id}")
        allocations: Counter[tuple[str, str]] = Counter()
        by_route: dict[str, sqlite3.Row] = {}
        for stratum in strata:
            by_route.setdefault(stratum["output_tier"], stratum)
        for route in ("Trusted", "Review", "Excluded"):
            if route in by_route:
                stratum = by_route[route]
                allocations[(stratum["output_tier"], stratum["primary_reason"])] += 1
        cursor = 0
        while sum(allocations.values()) < 13:
            stratum = strata[cursor % len(strata)]
            key = (stratum["output_tier"], stratum["primary_reason"])
            if allocations[key] < int(stratum["n"]):
                allocations[key] += 1
            cursor += 1
            if cursor > len(strata) * 20:
                raise RuntimeError(f"Unable to allocate 13 random sample rows for {output_id}")
        count_lookup = {(s["output_tier"], s["primary_reason"]): int(s["n"]) for s in strata}
        for (route, reason), allocation in sorted(allocations.items()):
            rows = connection.execute(
                f"""SELECT * FROM row_fact
                    WHERE run_id=? AND output_file_id=? AND output_tier=? AND primary_reason=?
                      AND row_fact_id NOT IN ({placeholders})
                    ORDER BY stable_rand(?,source_row_id) LIMIT ?""",
                (run_id, output_id, route, reason, *targeted_exclusions, fixed_seed, allocation),
            ).fetchall()
            population = count_lookup[(route, reason)]
            probability = allocation / population
            for row in rows:
                chosen.add(int(row["row_fact_id"]))
                sample_rank += 1
                connection.execute(insert_sql, (
                    row["row_fact_id"], run_id, output_id, row["source_row_id"],
                    "Deterministic stratified random", f"{route} | {reason}", "Random coverage",
                    probability, 1.0 / probability, fixed_seed, sample_rank,
                    f"Seeded without-replacement draw: stratum={route} | {reason}; n={population}; k={allocation}.",
                    shadow_recommendation(row),
                ))
        actual = connection.execute(
            "SELECT COUNT(*),SUM(sample_type='Targeted'),SUM(sample_type='Deterministic stratified random') "
            "FROM review_label WHERE run_id=? AND output_file_id=?",
            (run_id, output_id),
        ).fetchone()
        if tuple(actual) != (25, 12, 13):
            raise RuntimeError(f"Sample contract failed for {output_id}: {tuple(actual)}")
    connection.commit()


def build_funnel_cube(connection: sqlite3.Connection, run_id: str) -> None:
    connection.execute("DELETE FROM funnel_cube WHERE run_id=?", (run_id,))
    connection.execute("""
    INSERT INTO funnel_cube (
      run_id,output_file_id,output_label,funnel_type,stage_id,stage_order,stage_label,
      rule_version,candidate_status,outcome,transaction_count,value_usd,volume,
      missing_value_count,invalid_value_count,missing_volume_count,invalid_volume_count,
      weighted_asp,previous_stage_transaction_count,previous_stage_value_usd,
      previous_stage_volume,filtered_transaction_count,filtered_value_usd,filtered_volume,
      filtered_value_pct
    )
    WITH scoped AS (
      SELECT st.*, f.value_usd,f.volume,f.value_numeric_status,f.volume_numeric_status,
             s.output_label AS physical_label
      FROM row_stage_state st JOIN row_fact f ON f.row_fact_id=st.row_fact_id
      JOIN source_file s ON s.run_id=st.run_id AND s.output_file_id=st.output_file_id
      WHERE st.run_id=?
      UNION ALL
      SELECT st.row_fact_id,st.run_id,'OVERALL',st.source_row_id,st.stage_id,st.stage_order,
             st.rule_id,st.rule_version,st.candidate_status,st.outcome,st.reason,st.row_continues,
             f.value_usd,f.volume,f.value_numeric_status,f.volume_numeric_status,'Overall'
      FROM row_stage_state st JOIN row_fact f ON f.row_fact_id=st.row_fact_id
      WHERE st.run_id=?
    ), den AS (
      SELECT output_file_id,stage_id,COUNT(*) n,SUM(value_usd) value_usd,SUM(volume) volume
      FROM scoped GROUP BY output_file_id,stage_id
    ), agg AS (
      SELECT output_file_id,physical_label,stage_id,stage_order,rule_version,candidate_status,outcome,
             COUNT(*) n,SUM(value_usd) value_usd,SUM(volume) volume,
             SUM(value_numeric_status='missing') mv,SUM(value_numeric_status='invalid') iv,
             SUM(volume_numeric_status='missing') mq,SUM(volume_numeric_status='invalid') iq
      FROM scoped GROUP BY output_file_id,physical_label,stage_id,stage_order,rule_version,candidate_status,outcome
    )
    SELECT ?,a.output_file_id,a.physical_label,
           CASE rs.stage_type
             WHEN 'candidate' THEN 'Candidate'
             WHEN 'terminal' THEN 'Terminal'
             WHEN 'presentation' THEN 'Presentation'
             ELSE 'Extraction'
           END,
           a.stage_id,a.stage_order,rs.documentation_label,a.rule_version,a.candidate_status,a.outcome,
           a.n,a.value_usd,a.volume,a.mv,a.iv,a.mq,a.iq,
           CASE WHEN a.volume IS NULL OR abs(a.volume)<1e-18 THEN NULL ELSE a.value_usd/a.volume END,
           d.n,d.value_usd,d.volume,
           CASE WHEN a.candidate_status IN ('Suppressed','Missed') THEN a.n ELSE 0 END,
           CASE WHEN a.candidate_status IN ('Suppressed','Missed') THEN a.value_usd ELSE 0 END,
           CASE WHEN a.candidate_status IN ('Suppressed','Missed') THEN a.volume ELSE 0 END,
           CASE WHEN a.candidate_status IN ('Suppressed','Missed') AND d.value_usd<>0 THEN a.value_usd/d.value_usd ELSE 0 END
    FROM agg a JOIN den d USING(output_file_id,stage_id)
    JOIN rule_registry_stage rs ON rs.stage_id=a.stage_id
    ORDER BY a.output_file_id,a.stage_order,a.candidate_status,a.outcome
    """, (run_id, run_id, run_id))
    connection.execute("""
    INSERT INTO funnel_cube (
      run_id,output_file_id,output_label,funnel_type,stage_id,stage_order,stage_label,
      rule_version,candidate_status,outcome,transaction_count,value_usd,volume,
      missing_value_count,invalid_value_count,missing_volume_count,invalid_volume_count,
      weighted_asp,previous_stage_transaction_count,previous_stage_value_usd,
      previous_stage_volume,filtered_transaction_count,filtered_value_usd,filtered_volume,
      filtered_value_pct
    )
    WITH physical AS (
      SELECT s.run_id,s.output_file_id,s.output_label,s.transaction_count,s.value_usd,s.volume,
             s.missing_value_count,s.invalid_value_count,s.missing_volume_count,s.invalid_volume_count
      FROM source_file s WHERE s.run_id=?
    ), presented AS (
      SELECT * FROM physical
      UNION ALL
      SELECT run_id,'OVERALL','Overall',SUM(transaction_count),SUM(value_usd),SUM(volume),
             SUM(missing_value_count),SUM(invalid_value_count),
             SUM(missing_volume_count),SUM(invalid_volume_count)
      FROM physical GROUP BY run_id
    )
    SELECT p.run_id,p.output_file_id,p.output_label,'Presentation',s.stage_id,s.execution_order,
           s.documentation_label,r.version,'RetainedInSQLite','RetainedInSQLite',
           p.transaction_count,p.value_usd,p.volume,p.missing_value_count,p.invalid_value_count,
           p.missing_volume_count,p.invalid_volume_count,
           CASE WHEN p.volume IS NULL OR abs(p.volume)<1e-18 THEN NULL ELSE p.value_usd/p.volume END,
           p.transaction_count,p.value_usd,p.volume,0,0,0,0
    FROM presented p
    JOIN rule_registry_stage s ON s.stage_id='S14_PRESENTATION_EXPORT'
    JOIN rule_registry_rule r ON r.rule_id='S14_BOUNDED_PRESENTATION'
    """, (run_id,))
    connection.commit()


def normalized_dimension(column: str, active: bool) -> str:
    if not active:
        return f"'{DIMENSION_ALL}'"
    return f"CASE WHEN {column} IS NULL OR trim({column})='' THEN '{DIMENSION_UNMAPPED}' ELSE {column} END"


def build_removal_cube(connection: sqlite3.Connection, run_id: str) -> None:
    connection.execute("DELETE FROM removal_cube WHERE run_id=?", (run_id,))
    levels = [
        ("All", False, False, False),
        ("Manufacturer", True, False, False),
        ("Family", False, True, False),
        ("Product", False, False, True),
        ("Manufacturer × Family", True, True, False),
        ("Manufacturer × Family × Product", True, True, True),
    ]
    for level, use_m, use_f, use_p in levels:
        m = normalized_dimension("f.manufacturer", use_m)
        fam = normalized_dimension("f.family", use_f)
        prod = normalized_dimension("f.product", use_p)
        dm = normalized_dimension("df.manufacturer", use_m)
        dfam = normalized_dimension("df.family", use_f)
        dprod = normalized_dimension("df.product", use_p)
        connection.execute(f"""
        INSERT INTO removal_cube (
          run_id,output_file_id,output_label,reason_kind,is_additive,stage_id,rule_id,
          rule_version,outcome,reason,grouping_level,grouping_id,manufacturer,family,product,
          transaction_count,value_usd,volume,missing_value_count,invalid_value_count,
          missing_volume_count,invalid_volume_count,weighted_asp,
          previous_stage_transaction_count,previous_stage_value_usd,previous_stage_volume,
          filtered_transaction_count,filtered_value_usd,filtered_volume,filtered_value_pct
        )
        WITH hit_scope AS (
          SELECT h.*,f.output_tier,f.value_usd,f.volume,f.value_numeric_status,f.volume_numeric_status,
                 {m} manufacturer,{fam} family,{prod} product,s.output_label
          FROM rule_hit h JOIN row_fact f ON f.row_fact_id=h.row_fact_id
          JOIN source_file s ON s.run_id=h.run_id AND s.output_file_id=h.output_file_id
          WHERE h.run_id=?
          UNION ALL
          SELECT h.rule_hit_id,h.row_fact_id,h.run_id,'OVERALL',h.source_row_id,h.stage_id,h.rule_id,
                 h.rule_version,h.reason,h.hit_kind,h.is_additive,f.output_tier,f.value_usd,f.volume,
                 f.value_numeric_status,f.volume_numeric_status,{m},{fam},{prod},'Overall'
          FROM rule_hit h JOIN row_fact f ON f.row_fact_id=h.row_fact_id WHERE h.run_id=?
        ), den_scope AS (
          SELECT st.output_file_id,st.stage_id,{dm} manufacturer,{dfam} family,{dprod} product,
                 df.value_usd,df.volume
          FROM row_stage_state st JOIN row_fact df ON df.row_fact_id=st.row_fact_id WHERE st.run_id=?
          UNION ALL
          SELECT 'OVERALL',st.stage_id,{dm},{dfam},{dprod},df.value_usd,df.volume
          FROM row_stage_state st JOIN row_fact df ON df.row_fact_id=st.row_fact_id WHERE st.run_id=?
        ), den AS (
          SELECT output_file_id,stage_id,manufacturer,family,product,COUNT(*) n,
                 SUM(value_usd) value_usd,SUM(volume) volume
          FROM den_scope GROUP BY output_file_id,stage_id,manufacturer,family,product
        ), agg AS (
          SELECT output_file_id,output_label,hit_kind,is_additive,stage_id,rule_id,rule_version,
                 output_tier,reason,manufacturer,family,product,COUNT(*) n,
                 SUM(value_usd) value_usd,SUM(volume) volume,
                 SUM(value_numeric_status='missing') mv,SUM(value_numeric_status='invalid') iv,
                 SUM(volume_numeric_status='missing') mq,SUM(volume_numeric_status='invalid') iq
          FROM hit_scope
          GROUP BY output_file_id,output_label,hit_kind,is_additive,stage_id,rule_id,rule_version,
                   output_tier,reason,manufacturer,family,product
        )
        SELECT ?,a.output_file_id,a.output_label,
               CASE a.hit_kind WHEN 'primary' THEN 'Primary' ELSE 'Secondary' END,a.is_additive,
               a.stage_id,a.rule_id,a.rule_version,a.output_tier,a.reason,?,
               ? || '|' || a.manufacturer || '|' || a.family || '|' || a.product,
               a.manufacturer,a.family,a.product,a.n,a.value_usd,a.volume,a.mv,a.iv,a.mq,a.iq,
               CASE WHEN a.volume IS NULL OR abs(a.volume)<1e-18 THEN NULL ELSE a.value_usd/a.volume END,
               d.n,d.value_usd,d.volume,
               CASE WHEN a.hit_kind='primary' THEN a.n ELSE 0 END,
               CASE WHEN a.hit_kind='primary' THEN a.value_usd ELSE 0 END,
               CASE WHEN a.hit_kind='primary' THEN a.volume ELSE 0 END,
               CASE WHEN a.hit_kind='primary' AND d.value_usd<>0 THEN a.value_usd/d.value_usd ELSE 0 END
        FROM agg a LEFT JOIN den d ON d.output_file_id=a.output_file_id AND d.stage_id=a.stage_id
          AND d.manufacturer=a.manufacturer AND d.family=a.family AND d.product=a.product
        """, (run_id, run_id, run_id, run_id, run_id, level, level))
        connection.commit()
    connection.execute(
        "CREATE INDEX idx_removal_cube_lookup ON removal_cube(run_id,output_file_id,reason_kind,grouping_level,stage_id)"
    )
    connection.commit()


def add_qc(
    connection: sqlite3.Connection, run_id: str, output_id: str, check_name: str,
    observed: object, expected: object, passed: bool, evidence: str,
    value_delta: float | None = None, volume_delta: float | None = None,
    warn: bool = False,
) -> None:
    status = "WARN" if warn else ("PASS" if passed else "FAIL")
    connection.execute(
        """INSERT INTO reconciliation_qc
        (run_id,output_file_id,check_name,observed,expected,value_delta,volume_delta,status,evidence,checked_at)
        VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (run_id, output_id, check_name, str(observed), str(expected), value_delta, volume_delta,
         status, evidence, utc_now()),
    )


def build_baseline_manifest(
    connection: sqlite3.Connection, run_id: str, repo: Path, config_path: Path,
    registry_path: Path, sources: dict[str, object], code_path: Path,
) -> None:
    paths: list[tuple[str, Path, tuple[object, object, object] | None]] = [
        ("code", code_path, None), ("configuration", config_path, None),
        ("rule_registry", registry_path, None),
    ]
    for relative in sources.get("reference_sources", []):
        paths.append(("reference_source", repo / str(relative), None))
    source_metrics = {
        row[0]: (row[1], row[2], row[3]) for row in connection.execute(
            "SELECT output_file_id,transaction_count,value_usd,volume FROM source_file WHERE run_id=?",
            (run_id,),
        )
    }
    for spec in sources["outputs"]:
        source_path = repo / str(spec["path"])
        complete_source_path = (
            repo / str(spec["complete_source_path"])
            if spec.get("complete_source_path") else source_path
        )
        metrics = source_metrics.get(str(spec["output_file_id"]))
        paths.append(("mapped_complete_source", source_path, metrics))
        if complete_source_path.resolve() != source_path.resolve():
            paths.append(("immutable_complete_input", complete_source_path, None))
    for kind, path, metrics in paths:
        if not path.exists():
            raise FileNotFoundError(path)
        transaction_count, value, volume = metrics or (None, None, None)
        connection.execute(
            """INSERT OR REPLACE INTO baseline_manifest
            (run_id,artifact_type,path,sha256,bytes,transaction_count,value_usd,volume)
            VALUES (?,?,?,?,?,?,?,?)""",
            (run_id, kind, str(path.resolve()), sha256_file(path), path.stat().st_size,
             transaction_count, value, volume),
        )
    connection.commit()


def build_reconciliation_qc(
    connection: sqlite3.Connection, run_id: str, expected_outputs: Sequence[str]
) -> None:
    connection.execute("DELETE FROM reconciliation_qc WHERE run_id=?", (run_id,))
    for output_id in expected_outputs:
        source = connection.execute(
            """SELECT output_label,expected_rows,observed_rows,transaction_count,value_usd,volume,
                      is_complete,ingestion_mode FROM source_file
               WHERE run_id=? AND output_file_id=?""", (run_id, output_id)
        ).fetchone()
        if source is None:
            add_qc(connection, run_id, output_id, "Source registered", 0, 1, False, "Missing source_file row")
            continue
        label, expected, observed, count, value, volume, complete, mode = source
        add_qc(connection, run_id, output_id, "Complete source row count", observed, expected,
               observed == expected == count, f"{label}; ingestion_mode={mode}; is_complete={complete}")
        for stage_id in ("S00_EXTRACTION", "S13_TERMINAL_ROUTING"):
            stage_count = connection.execute(
                "SELECT COUNT(*) FROM row_stage_state WHERE run_id=? AND output_file_id=? AND stage_id=?",
                (run_id, output_id, stage_id),
            ).fetchone()[0]
            add_qc(connection, run_id, output_id, f"Exactly one {stage_id} state per row",
                   stage_count, count, stage_count == count, "Normalized stage-state cardinality")
        terminal = connection.execute(
            f"SELECT {metrics_sql()} FROM row_fact WHERE run_id=? AND output_file_id=?",
            (run_id, output_id),
        ).fetchone()
        value_delta = float(terminal[1] or 0) - float(value or 0)
        volume_delta = float(terminal[2] or 0) - float(volume or 0)
        passed = terminal[0] == count and abs(value_delta) <= VALUE_TOLERANCE and abs(volume_delta) <= VOLUME_TOLERANCE
        add_qc(connection, run_id, output_id, "Source-to-terminal totals", terminal[:3], (count, value, volume),
               passed, "Transaction/value/volume reconciliation", value_delta, volume_delta)
        sample = connection.execute(
            "SELECT COUNT(*),SUM(sample_type='Targeted'),SUM(sample_type='Deterministic stratified random') "
            "FROM review_label WHERE run_id=? AND output_file_id=?", (run_id, output_id)
        ).fetchone()
        add_qc(connection, run_id, output_id, "Review sample contract", tuple(sample), (25, 12, 13),
               tuple(sample) == (25, 12, 13), "Exact purposeful + deterministic-stratified design")
        route_total = connection.execute(
            "SELECT COUNT(*) FROM row_fact WHERE run_id=? AND output_file_id=? AND output_tier IN ('Trusted','Review','Excluded')",
            (run_id, output_id),
        ).fetchone()[0]
        add_qc(connection, run_id, output_id, "Exactly one terminal route", route_total, count,
               route_total == count, "Terminal route is constrained in row_fact")
        funnel_terminal = connection.execute(
            "SELECT SUM(transaction_count) FROM funnel_cube WHERE run_id=? AND output_file_id=? AND stage_id='S13_TERMINAL_ROUTING'",
            (run_id, output_id),
        ).fetchone()[0] or 0
        add_qc(connection, run_id, output_id, "Funnel terminal reconciliation", funnel_terminal, count,
               funnel_terminal == count, "S13 funnel partitions the physical output")
        if output_id == "IN_2025":
            passed_india = count > EXCEL_MAX_DATA_ROWS and mode == "complete_csv_current_remap" and complete == 1
            add_qc(connection, run_id, output_id, "India FY2025 no Excel-cap truncation", count,
                   f">{EXCEL_MAX_DATA_ROWS} complete_csv_current_remap", passed_india,
                   "Complete CSV is authoritative; capped workbook is rejected")
        if output_id.startswith("PK_"):
            bad_nonstandard = connection.execute(
                "SELECT COUNT(*) FROM row_fact WHERE run_id=? AND output_file_id=? AND nonstandard_tier=1 AND output_tier='Trusted'",
                (run_id, output_id),
            ).fetchone()[0]
            add_qc(connection, run_id, output_id, "Pakistan nonstandard tiers not Trusted", bad_nonstandard, 0,
                   bad_nonstandard == 0, "Manufacturer/HS-prior tiers remain visible and reviewable")

    source_count = connection.execute(
        "SELECT COUNT(*) FROM source_file WHERE run_id=?", (run_id,)
    ).fetchone()[0]
    add_qc(connection, run_id, "OVERALL", "Six governed outputs present", source_count, 6,
           source_count == 6, "Required country-year output scope")
    overall_count = connection.execute(
        "SELECT SUM(transaction_count) FROM source_file WHERE run_id=?", (run_id,)
    ).fetchone()[0]
    funnel_overall = connection.execute(
        "SELECT SUM(transaction_count) FROM funnel_cube WHERE run_id=? AND output_file_id='OVERALL' AND stage_id='S13_TERMINAL_ROUTING'",
        (run_id,),
    ).fetchone()[0] or 0
    add_qc(connection, run_id, "OVERALL", "Overall funnel reconciliation", funnel_overall, overall_count,
           funnel_overall == overall_count, "Overall duplicates each physical row exactly once")
    mri_fact = connection.execute("SELECT COUNT(*) FROM row_fact WHERE run_id=? AND mri_risk=1", (run_id,)).fetchone()[0]
    mri_inventory = connection.execute(
        "SELECT COUNT(*) FROM recall_risk_inventory WHERE run_id=? AND risk_type='MRI compatible recall risk'", (run_id,)
    ).fetchone()[0]
    add_qc(connection, run_id, "OVERALL", "Complete MRI-compatible inventory", mri_inventory, mri_fact,
           mri_inventory == mri_fact, "Risk inventory is complete, not sampled")
    imaging_overlap = connection.execute(
        "SELECT COUNT(*) FROM row_fact WHERE run_id=? AND mri_actual_imaging=1 AND mri_perioperative_signal=1", (run_id,)
    ).fetchone()[0]
    add_qc(connection, run_id, "OVERALL", "MRI imaging and perioperative flags retained separately",
           imaging_overlap, ">=0", True, "Overlap is allowed; both explicit flags remain queryable")
    invalid_grouping = connection.execute(
        """SELECT COUNT(*) FROM removal_cube WHERE run_id=? AND (
             grouping_level NOT IN ('All','Manufacturer','Family','Product','Manufacturer × Family','Manufacturer × Family × Product')
             OR grouping_id IS NULL OR manufacturer IS NULL OR family IS NULL OR product IS NULL)""", (run_id,)
    ).fetchone()[0]
    add_qc(connection, run_id, "OVERALL", "Removal grouping schema", invalid_grouping, 0,
           invalid_grouping == 0, "Six explicit grouping levels with stable sentinel values")
    all_overall_rows = connection.execute(
        "SELECT COUNT(*) FROM removal_cube WHERE run_id=? AND output_file_id='OVERALL' AND grouping_level='All'",
        (run_id,),
    ).fetchone()[0]
    add_qc(connection, run_id, "OVERALL", "Overall removal cube present", all_overall_rows, ">0",
           all_overall_rows > 0, "Overall plus six physical outputs are retained")
    production_changes = connection.execute(
        "SELECT COUNT(*) FROM review_label WHERE run_id=? AND production_changed<>0", (run_id,)
    ).fetchone()[0]
    add_qc(connection, run_id, "OVERALL", "Review-only boundary", production_changes, 0,
           production_changes == 0, "All recommendations are shadow-only")
    foreign_keys = connection.execute("PRAGMA foreign_key_check").fetchall()
    add_qc(connection, run_id, "OVERALL", "SQLite foreign keys", len(foreign_keys), 0,
           not foreign_keys, "PRAGMA foreign_key_check")
    integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
    add_qc(connection, run_id, "OVERALL", "SQLite integrity", integrity, "ok",
           integrity == "ok", "PRAGMA integrity_check")
    connection.commit()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the complete review-only prediction audit authority.")
    parser.add_argument("--sources", default="config/audit_sources.json")
    parser.add_argument("--registry", default="config/prediction_rule_registry.json")
    parser.add_argument("--output", default=None, help="Override the final SQLite output path.")
    parser.add_argument(
        "--resume-building",
        action="store_true",
        help="Resume aggregate construction from a validated .building database after a post-ingestion failure.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo = Path(__file__).resolve().parents[1]
    config_path = (repo / args.sources).resolve()
    registry_path = (repo / args.registry).resolve()
    sources = json.loads(config_path.read_text(encoding="utf-8"))
    registry = json.loads(registry_path.read_text(encoding="utf-8"))
    validate_registry(registry)
    run_id = str(sources["run_id"])
    fixed_seed = int(sources.get("fixed_seed", registry.get("fixed_seed", 20260710)))
    outputs = sources.get("outputs", [])
    output_ids = [str(item["output_file_id"]) for item in outputs]
    required_ids = ["VN_2024", "VN_2025", "PK_2024", "PK_2025", "IN_2024", "IN_2025"]
    if output_ids != required_ids:
        raise RuntimeError(f"FAIL CLOSED: configured output order must be {required_ids}; observed {output_ids}")
    final_path = Path(args.output).resolve() if args.output else repo / "outputs" / run_id / "prediction_audit.sqlite"
    final_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = final_path.with_suffix(final_path.suffix + ".building")
    connection: sqlite3.Connection | None = None
    try:
        if args.resume_building:
            if not temp_path.exists():
                raise RuntimeError(f"No retained building database exists at {temp_path}")
            connection = sqlite3.connect(temp_path)
            connection.execute("PRAGMA foreign_keys=ON")
            connection.execute("PRAGMA journal_mode=DELETE")
            connection.execute("PRAGMA synchronous=NORMAL")
            connection.execute("PRAGMA temp_store=FILE")
            connection.execute("PRAGMA cache_size=-250000")
            existing = connection.execute(
                "SELECT registry_version,fixed_seed FROM run WHERE run_id=?", (run_id,)
            ).fetchone()
            expected_counts = {str(spec["output_file_id"]): int(spec["expected_rows"]) for spec in outputs}
            observed_counts = dict(connection.execute(
                "SELECT output_file_id,observed_rows FROM source_file WHERE run_id=?", (run_id,)
            ).fetchall())
            if existing != (registry["registry_version"], fixed_seed) or observed_counts != expected_counts:
                raise RuntimeError(
                    "FAIL CLOSED: retained building database does not match the registry, seed, and configured source counts"
                )
            connection.execute(
                "UPDATE run SET completed_at=NULL,status='building' WHERE run_id=?", (run_id,)
            )
            connection.commit()
            print(f"Resuming validated post-ingestion database {temp_path}...", flush=True)
        else:
            connection = initialize_db(temp_path)
            connection.execute(
                "INSERT INTO run VALUES (?,?,?,?,?,?,?,?,?)",
                (run_id, utc_now(), None, registry["registry_version"], fixed_seed, "review_only",
                 "building", git_commit(), "Complete-source audit authority; no production routing changes."),
            )
            connection.commit()
            rules = load_registry(connection, registry)
            for spec in outputs:
                print(f"Loading {spec['output_label']}...", flush=True)
                ingest_output(connection, run_id, spec, rules)
            create_indexes_and_views(connection)
        print("Building deterministic review samples...", flush=True)
        build_review_samples(connection, run_id, fixed_seed)
        print("Building funnel cube...", flush=True)
        build_funnel_cube(connection, run_id)
        print("Building removal cube...", flush=True)
        build_removal_cube(connection, run_id)
        build_baseline_manifest(connection, run_id, repo, config_path, registry_path, sources, Path(__file__).resolve())
        build_reconciliation_qc(connection, run_id, required_ids)
        failed = connection.execute(
            "SELECT COUNT(*) FROM reconciliation_qc WHERE run_id=? AND status='FAIL'", (run_id,)
        ).fetchone()[0]
        if failed:
            raise RuntimeError(f"FAIL CLOSED: {failed} reconciliation checks failed")
        connection.execute(
            "UPDATE run SET completed_at=?,status='passed' WHERE run_id=?", (utc_now(), run_id)
        )
        connection.execute("ANALYZE")
        connection.execute("PRAGMA optimize")
        connection.commit()
        connection.close()
        connection = None
        os.replace(temp_path, final_path)
        print(f"Published {final_path}", flush=True)
        return 0
    except Exception:
        if connection is not None:
            try:
                connection.execute("UPDATE run SET completed_at=?,status='failed' WHERE run_id=?", (utc_now(), run_id))
                connection.commit()
            except Exception:
                pass
            connection.close()
        raise


if __name__ == "__main__":
    raise SystemExit(main())
