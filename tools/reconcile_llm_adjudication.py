"""Reconcile LLM reviews in SQLite and export governed final workbooks.

SQLite remains the processing authority.  Excel files are presentation/final
handoff artifacts only and contain no formulas or hidden processing state.
"""

from __future__ import annotations

import argparse
import json
import sqlite3
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RUN = ROOT / "outputs" / "20260713_llm_adjudication"
BLUE = "0047AB"
LIGHT_BLUE = "DCE8F8"
LIGHT_GRAY = "E7E6E6"


def _json(value: str) -> dict:
    return json.loads(value)


def _style_sheet(ws, widths: dict[str, int] | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    for idx, cell in enumerate(ws[1], 1):
        width = (widths or {}).get(str(cell.value), 18)
        ws.column_dimensions[get_column_letter(idx)].width = min(width, 60)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_rows(ws, columns: list[str], rows: list[dict]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])


def _proposal_reconciliation(conn: sqlite3.Connection, threshold: float) -> list[dict]:
    conn.executescript(
        """
        DROP TABLE IF EXISTS proposal_final;
        CREATE TABLE proposal_final (
            proposal_id TEXT PRIMARY KEY REFERENCES proposal_raw(proposal_id),
            first_reviewer TEXT NOT NULL,
            first_decision TEXT NOT NULL,
            first_confidence REAL NOT NULL,
            cross_reviewer TEXT,
            cross_decision TEXT,
            cross_confidence REAL,
            final_decision TEXT NOT NULL,
            final_confidence REAL NOT NULL,
            final_rationale TEXT NOT NULL,
            reconciled_at_utc TEXT NOT NULL
        );
        """
    )
    raw = conn.execute(
        "SELECT proposal_id, payload_json FROM proposal_raw ORDER BY source_excel_row"
    ).fetchall()
    result = []
    now = datetime.now(timezone.utc).isoformat()
    for proposal_id, payload_json in raw:
        reviews = conn.execute(
            "SELECT reviewer,decision,confidence,rationale FROM proposal_review WHERE proposal_id=?",
            (proposal_id,),
        ).fetchall()
        first = [r for r in reviews if "cross-pass" not in r[0].lower()]
        cross = [r for r in reviews if "proposal cross-pass" in r[0].lower()]
        if len(first) != 1:
            raise SystemExit(f"{proposal_id}: expected one first review, found {len(first)}")
        first_reviewer, first_decision, first_confidence, first_rationale = first[0]
        cross_review = cross[0] if cross else ("", "", None, "")
        cross_reviewer, cross_decision, cross_confidence, cross_rationale = cross_review

        is_candidate = first_decision == "APPROVE" and first_confidence >= threshold
        if is_candidate and len(cross) != 1:
            raise SystemExit(f"{proposal_id}: second review missing for approval candidate")
        if is_candidate and cross_decision == "APPROVE" and cross_confidence >= threshold:
            final_decision = "APPROVE"
            final_confidence = min(first_confidence, cross_confidence)
        elif first_decision == "REJECT" or (is_candidate and cross_decision == "REJECT"):
            final_decision = "REJECT"
            final_confidence = max(first_confidence, cross_confidence or 0)
        else:
            final_decision = "HOLD"
            final_confidence = min(first_confidence, cross_confidence or first_confidence)
        rationale_parts = [f"First: {first_rationale}"]
        if cross_rationale:
            rationale_parts.append(f"Second: {cross_rationale}")
        final_rationale = " | ".join(rationale_parts)
        conn.execute(
            "INSERT INTO proposal_final VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (
                proposal_id, first_reviewer, first_decision, first_confidence,
                cross_reviewer, cross_decision, cross_confidence,
                final_decision, final_confidence, final_rationale, now,
            ),
        )
        payload = _json(payload_json)
        payload.update(
            {
                "First_Reviewer": first_reviewer,
                "First_Decision": first_decision,
                "First_Confidence": first_confidence,
                "Cross_Reviewer": cross_reviewer,
                "Cross_Decision": cross_decision,
                "Cross_Confidence": cross_confidence if cross_confidence is not None else "",
                "Final_Decision": final_decision,
                "Final_Confidence": final_confidence,
                "Final_Rationale": final_rationale,
            }
        )
        payload["Approved"] = "Y" if final_decision == "APPROVE" else ""
        payload["Reviewer_Notes"] = (
            "Two-reviewer LLM evidence panel, user-authorized 2026-07-13. "
            + final_rationale
        )
        result.append(payload)
    return result


def _precision_consensus(conn: sqlite3.Connection) -> list[dict]:
    conn.executescript(
        """
        DROP TABLE IF EXISTS precision_consensus;
        CREATE TABLE precision_consensus (
            precision_id TEXT PRIMARY KEY REFERENCES precision_raw(precision_id),
            reviewer_1 TEXT NOT NULL,
            relevance_1 TEXT NOT NULL,
            correctness_1 TEXT,
            reviewer_2 TEXT NOT NULL,
            relevance_2 TEXT NOT NULL,
            correctness_2 TEXT,
            consensus_relevance TEXT NOT NULL,
            consensus_correctness TEXT,
            exact_agreement INTEGER NOT NULL,
            consensus_rationale TEXT NOT NULL
        );
        """
    )
    raw = conn.execute(
        "SELECT precision_id,payload_json FROM precision_raw ORDER BY source_excel_row"
    ).fetchall()
    result = []
    for precision_id, payload_json in raw:
        reviews = conn.execute(
            """SELECT reviewer,surgical_relevance,mapping_correctness,reviewer_rationale
               FROM precision_review WHERE precision_id=? ORDER BY reviewer""",
            (precision_id,),
        ).fetchall()
        if len(reviews) != 2:
            raise SystemExit(f"{precision_id}: expected two precision reviews, found {len(reviews)}")
        r1, r2 = reviews
        consensus_relevance = r1[1] if r1[1] == r2[1] else "Uncertain"
        if consensus_relevance == "Surgical":
            consensus_correctness = r1[2] if r1[2] == r2[2] else "Uncertain"
        else:
            consensus_correctness = ""
        exact = int(r1[1] == r2[1] and r1[2] == r2[2])
        rationale = f"{r1[0]}: {r1[3]} | {r2[0]}: {r2[3]}"
        conn.execute(
            "INSERT INTO precision_consensus VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (
                precision_id, r1[0], r1[1], r1[2], r2[0], r2[1], r2[2],
                consensus_relevance, consensus_correctness, exact, rationale,
            ),
        )
        payload = _json(payload_json)
        payload.update(
            {
                "LLM_Reviewer_1": r1[0],
                "LLM_Relevance_1": r1[1],
                "LLM_Correctness_1": r1[2],
                "LLM_Reviewer_2": r2[0],
                "LLM_Relevance_2": r2[1],
                "LLM_Correctness_2": r2[2],
                "LLM_Consensus_Relevance": consensus_relevance,
                "LLM_Consensus_Correctness": consensus_correctness,
                "LLM_Exact_Agreement": exact,
                "LLM_Consensus_Rationale": rationale,
            }
        )
        result.append(payload)
    return result


def _ratio(numerator: float, denominator: float) -> float | None:
    return numerator / denominator if denominator else None


def _scores(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        """
        SELECT r.sample_type,r.output_tier,r.country,r.fiscal_year,r.sample_weight,
               r.value_usd,r.volume,c.consensus_relevance,c.consensus_correctness,
               c.exact_agreement
        FROM precision_raw r JOIN precision_consensus c USING (precision_id)
        """
    ).fetchall()
    scopes: dict[tuple[str, str], list] = defaultdict(list)
    for row in rows:
        sample_type, tier = row[0], row[1]
        scopes[(sample_type, "All tiers")].append(row)
        scopes[(sample_type, tier)].append(row)
    result = []
    for (sample_type, tier), group in sorted(scopes.items()):
        random = sample_type == "Deterministic stratified random"
        weights = [float(r[4] or 0) if random else 1.0 for r in group]
        determinate_rel = [(r, w) for r, w in zip(group, weights) if r[7] != "Uncertain"]
        surgical = [(r, w) for r, w in determinate_rel if r[7] == "Surgical"]
        determinate_map = [(r, w) for r, w in surgical if r[8] in {"Correct", "Incorrect"}]
        correct = [(r, w) for r, w in determinate_map if r[8] == "Correct"]
        rel_share = _ratio(sum(w for _, w in surgical), sum(w for _, w in determinate_rel))
        map_share = _ratio(sum(w for _, w in correct), sum(w for _, w in determinate_map))
        value_share = _ratio(
            sum(w * float(r[5] or 0) for r, w in correct),
            sum(w * float(r[5] or 0) for r, w in determinate_map),
        )
        volume_share = _ratio(
            sum(w * float(r[6] or 0) for r, w in correct),
            sum(w * float(r[6] or 0) for r, w in determinate_map),
        )
        result.append(
            {
                "sample_type": sample_type,
                "scope": tier,
                "sample_rows": len(group),
                "exact_agreement_pct": _ratio(sum(r[9] for r in group), len(group)),
                "determinate_relevance_rows": len(determinate_rel),
                "estimated_surgical_share": rel_share,
                "determinate_mapping_rows": len(determinate_map),
                "estimated_mapping_precision_rows": map_share,
                "estimated_mapping_precision_value": value_share,
                "estimated_mapping_precision_volume": volume_share,
                "status": "LLM estimate; not human-verified ground truth",
            }
        )
    return result


def _aggregate_json(conn: sqlite3.Connection, scores: list[dict]) -> dict:
    proposal_summary = [
        dict(zip(["proposal_type", "final_decision", "rows", "value_usd"], row))
        for row in conn.execute(
            """SELECT p.proposal_type,f.final_decision,COUNT(*),SUM(p.cluster_value_usd)
               FROM proposal_raw p JOIN proposal_final f USING(proposal_id)
               GROUP BY p.proposal_type,f.final_decision ORDER BY 1,2"""
        )
    ]
    market_summary = [
        dict(zip(["market", "fiscal_year", "final_decision", "rows", "value_usd"], row))
        for row in conn.execute(
            """SELECT p.market,p.fiscal_year,f.final_decision,COUNT(*),SUM(p.cluster_value_usd)
               FROM proposal_raw p JOIN proposal_final f USING(proposal_id)
               GROUP BY p.market,p.fiscal_year,f.final_decision ORDER BY 1,2,3"""
        )
    ]
    totals = conn.execute(
        """SELECT COUNT(*),SUM(CASE WHEN final_decision='APPROVE' THEN 1 ELSE 0 END),
                  SUM(CASE WHEN final_decision='APPROVE' THEN cluster_value_usd ELSE 0 END)
           FROM proposal_raw JOIN proposal_final USING(proposal_id)"""
    ).fetchone()
    return {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "processing_authority": "SQLite",
        "row_level_data_location": "raw_outputs/LLM_Review_Raw_Output.xlsx",
        "proposal_totals": {
            "reviewed": totals[0], "approved": totals[1], "approved_value_usd": totals[2]
        },
        "proposal_summary": proposal_summary,
        "market_summary": market_summary,
        "precision_scores": scores,
        "precision_caveat": "LLM estimates are decision support, not analyst-labelled ground truth.",
    }


def _save_raw_workbook(path: Path, proposal_rows: list[dict], precision_rows: list[dict], scores: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposal_Raw_All"
    proposal_columns = list(proposal_rows[0])
    _write_rows(ws, proposal_columns, proposal_rows)
    _style_sheet(ws, {"Evidence_Quote": 50, "Final_Rationale": 60, "Reviewer_Notes": 60})

    ws = wb.create_sheet("Precision_Raw_All")
    precision_columns = list(precision_rows[0])
    _write_rows(ws, precision_columns, precision_rows)
    _style_sheet(ws, {"Detailed Product": 55, "LLM_Consensus_Rationale": 60})

    ws = wb.create_sheet("Aggregate_Score_Check")
    score_columns = list(scores[0])
    _write_rows(ws, score_columns, scores)
    _style_sheet(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                cell.number_format = "0.0%"

    ws = wb.create_sheet("Methodology")
    method_rows = [
        ("Processing backend", "SQLite; Excel is final output only."),
        ("Proposal approval", "Two independent LLM reviewers must both APPROVE at confidence >= 0.90; any disagreement is held or rejected."),
        ("Precision labels", "Two independent LLM readings reconciled conservatively. These are estimates, not human-verified ground truth."),
        ("Random sample", "Uses the stored design weights. Targeted examples are reported separately and are not population estimates."),
        ("Value/volume score", "Among determinate, consensus-surgical rows, the share assessed mapping-correct weighted by import value or volume."),
        ("Governance", "Business analysts should audit the raw rows and can reverse or amend governed reference entries through adjudication."),
    ]
    ws.append(["Item", "Plain-language explanation"])
    for row in method_rows:
        ws.append(row)
    _style_sheet(ws, {"Item": 26, "Plain-language explanation": 100})
    wb.save(path)


def _save_approved_workbook(path: Path, proposal_rows: list[dict]) -> None:
    # Preserve the proposal workbook contract; append panel metadata after source columns.
    source_columns = [k for k in proposal_rows[0] if not k.startswith(("First_", "Cross_", "Final_"))]
    metadata = [
        "First_Reviewer", "First_Decision", "First_Confidence",
        "Cross_Reviewer", "Cross_Decision", "Cross_Confidence",
        "Final_Decision", "Final_Confidence", "Final_Rationale",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Recovery_Proposals"
    _write_rows(ws, source_columns + metadata, proposal_rows)
    _style_sheet(ws, {"Evidence_Quote": 50, "Reviewer_Notes": 60, "Final_Rationale": 60})
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-dir", type=Path, default=DEFAULT_RUN)
    parser.add_argument("--threshold", type=float, default=0.90)
    args = parser.parse_args()
    raw_dir = args.run_dir / "raw_outputs"
    dashboard_dir = args.run_dir / "dashboard"
    raw_dir.mkdir(parents=True, exist_ok=True)
    dashboard_dir.mkdir(parents=True, exist_ok=True)
    db = raw_dir / "llm_review_authority.sqlite"
    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys=ON")

    proposal_rows = _proposal_reconciliation(conn, args.threshold)
    precision_rows = _precision_consensus(conn)
    scores = _scores(conn)
    aggregate = _aggregate_json(conn, scores)
    conn.commit()

    raw_book = raw_dir / "LLM_Review_Raw_Output.xlsx"
    approval_book = raw_dir / "Recall_Recovery_Proposals_LLM_Reviewed.xlsx"
    _save_raw_workbook(raw_book, proposal_rows, precision_rows, scores)
    _save_approved_workbook(approval_book, proposal_rows)
    (dashboard_dir / "aggregate_data.json").write_text(
        json.dumps(aggregate, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    conn.close()
    print(json.dumps({
        "database": str(db),
        "raw_workbook": str(raw_book),
        "approval_workbook": str(approval_book),
        "aggregate_json": str(dashboard_dir / "aggregate_data.json"),
        "approved": aggregate["proposal_totals"]["approved"],
        "approved_value_usd": aggregate["proposal_totals"]["approved_value_usd"],
    }, indent=2))


if __name__ == "__main__":
    main()
